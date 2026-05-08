"""
Wind3 Reload Dashboard — Modulo "Genera Chiusure"
Autore: Salvo (Ambassador A6 — Campania & Puglia)

Questo modulo aggiunge una pagina alla dashboard che permette di generare
chiusure mensili personalizzate (Excel o PDF) per ogni Sales Account/District
Manager della zona.

Features:
- Selezione SA/DM (multiselect + opzione "Tutti")
- Selezione mesi (checkbox per ogni mese disponibile)
- Configurazione colonne con flag on/off per ogni metrica
- Fasce di prezzo (6 fasce, toggle singolo)
- Sezione moltiplicatori compenso (Franchising X2/X3/X4 e Dealer X2/X3/X4)
- Output Excel (.xlsx) o PDF (landscape A4)
- Anteprima Streamlit dei dati esportati
- ZIP con tutti i file generati

Integrazione in app.py:
1. Aggiungi in cima: from chiusure import render_chiusure_tab
2. Aggiungi un nuovo tab nella lista st.tabs(...)
3. In quel tab, chiama: render_chiusure_tab(file_bytes, zona_label, regioni, sheet_name, filter_mode)
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ─────────────────────────────────────────────────────────────────
#  COSTANTI: COLORI, MOLTIPLICATORI, MAPPING CANALI
# ─────────────────────────────────────────────────────────────────
HEADER_HEX = "1B4F91"          # header principale
SUBHEADER_HEX = "2E75B6"       # subheader / aprile
NET_HEADER_HEX = "375623"      # header NET / NET% (verde scuro, sempre in evidenza)

# Palette per separare visivamente i mesi: alterna due tonalità di blu
MONTH_COLORS_HEX = ["1F4E79", "2E75B6", "4A86C5", "1F4E79", "2E75B6", "4A86C5"]

# Semaforo AR% e NET%
AR_GREEN_HEX = "C8E6C9"
AR_YELLOW_HEX = "FFF3C4"
AR_RED_HEX = "FFCDD2"

# Semaforo Optout%
OPTOUT_GREEN_HEX = "C8E6C9"
OPTOUT_YELLOW_HEX = "FFF3C4"
OPTOUT_RED_HEX = "FFCDD2"

# Forever (highlight binario)
FOREVER_GREEN_HEX = "C8E6C9"
FOREVER_RED_HEX = "FFCDD2"

# Compenso tier raggiunto
TIER_REACHED_HEX = "A5D6A7"

# Riga totale
TOTAL_ROW_HEX = "D9E5F4"

# Mesi italiani
MESI_IT = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
]

# Soglie AR% per moltiplicatore (richieste Salvo per chiusure)
MULTIPLIER_THRESHOLDS = {
    "Franchising": {"x2": 40.0, "x3": 60.0, "x4": 75.0},
    "Dealer":      {"x2": 25.0, "x3": 40.0, "x4": 55.0},
}

# Compenso a pezzo (Net × 5€)
COMPENSO_PER_PEZZO = 5.0

# Mapping STORE_TYPE → canale per moltiplicatori
# Large Chain → Dealer (specifica Salvo)
STORE_TYPE_TO_CHANNEL = {
    "Franchising":  "Franchising",
    "Dealer":       "Dealer",
    "Large Chain":  "Dealer",
    "Owned Stores": "Dealer",
    "Other":        "Dealer",
}

# Fasce prezzo (mappatura esattamente come nelle 76 colonne del file Wind3)
FASCE_DEF = [
    {"key": "0_150",     "label": "0-150€",      "tam_col": "TAM_0_150",      "gross_col": "SR_PLUS_GROSS_SALES_0_150"},
    {"key": "150_300",   "label": "150-300€",    "tam_col": "TAM_150_300",    "gross_col": "SR_PLUS_GROSS_SALES_150_300"},
    {"key": "300_700",   "label": "300-700€",    "tam_col": "TAM_300_700",    "gross_col": "SR_PLUS_GROSS_SALES_300_700"},
    {"key": "700_1200",  "label": "700-1200€",   "tam_col": "TAM_700_1200",   "gross_col": "SR_PLUS_GROSS_SALES_700_1200"},
    {"key": "1200_1600", "label": "1200-1600€",  "tam_col": "TAM_1200_1600",  "gross_col": "SR_PLUS_GROSS_SALES_1200_1600"},
    {"key": ">1600",     "label": ">1600€",      "tam_col": "TAM_>1600",      "gross_col": "SR_PLUS_GROSS_SALES_>1600"},
]

# Metriche base configurabili
# (codice interno, label visualizzata, colonne sorgente / formula)
METRICHE_BASE = [
    {"key": "qta",         "label": "Quantità",            "default": True},
    {"key": "gross",       "label": "Gross Sales",         "default": True},
    {"key": "ar_gross",    "label": "AR% Gross",           "default": True},
    {"key": "net",         "label": "Net Sales",           "default": True},
    {"key": "ar_net",      "label": "NET%",                "default": True, "highlight": True},
    {"key": "optout",      "label": "Optout",              "default": False},
    {"key": "optout_pct",  "label": "Optout%",             "default": False},
    {"key": "fb_gross",    "label": "Forever Basic Gross", "default": False},
    {"key": "fb_net",      "label": "Forever Basic Net",   "default": False},
    {"key": "fp_gross",    "label": "Forever Premium Gross","default": False},
    {"key": "fp_net",      "label": "Forever Premium Net", "default": False},
    {"key": "exchange",    "label": "Exchange",            "default": False},
]


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────
def _safe_div(num, den) -> float:
    """Divisione protetta che restituisce 0 in caso di zero/None/NaN."""
    try:
        if den is None or pd.isna(den) or float(den) == 0:
            return 0.0
        return float(num) / float(den)
    except Exception:
        return 0.0


def _safe_pct(num, den) -> float:
    """Percentuale 0-100 con 1 decimale."""
    return round(_safe_div(num, den) * 100.0, 1)


def _safe_str(v, default: str = "") -> str:
    if v is None:
        return default
    try:
        if pd.isna(v):
            return default
    except Exception:
        pass
    return str(v)


def _safe_int(v, default: int = 0) -> int:
    try:
        if v is None or pd.isna(v):
            return default
        return int(float(v))
    except Exception:
        return default


def _safe_float(v, default: float = 0.0) -> float:
    try:
        if v is None or pd.isna(v):
            return default
        return float(v)
    except Exception:
        return default


def _month_label(m: int, y: int) -> str:
    """Es. (4, 2026) → 'Aprile 2026'."""
    if 1 <= m <= 12:
        return f"{MESI_IT[m - 1]} {y}"
    return f"{m:02d}/{y}"


def _channel_for_store_type(store_type: str) -> str:
    return STORE_TYPE_TO_CHANNEL.get(_safe_str(store_type), "Dealer")


def _multiplier_from_ar(ar_net_pct: float, channel: str) -> int:
    """Restituisce il moltiplicatore X1/X2/X3/X4 raggiunto dato l'AR Net% e il canale."""
    s = MULTIPLIER_THRESHOLDS.get(channel, MULTIPLIER_THRESHOLDS["Dealer"])
    if ar_net_pct >= s["x4"]:
        return 4
    if ar_net_pct >= s["x3"]:
        return 3
    if ar_net_pct >= s["x2"]:
        return 2
    return 1


def _color_ar_hex(pct: float) -> str | None:
    """Verde ≥70%, giallo 40-70%, rosso <40%."""
    try:
        v = float(pct)
    except Exception:
        return None
    if v < 40:
        return AR_RED_HEX
    if v < 70:
        return AR_YELLOW_HEX
    return AR_GREEN_HEX


def _color_optout_hex(pct: float) -> str | None:
    """Verde ≤20%, giallo 20-50%, rosso >50%."""
    try:
        v = float(pct)
    except Exception:
        return None
    if v > 50:
        return OPTOUT_RED_HEX
    if v > 20:
        return OPTOUT_YELLOW_HEX
    return OPTOUT_GREEN_HEX


def _color_forever_hex(val: float) -> str | None:
    """Verde se >0, rosso se =0."""
    try:
        v = float(val)
    except Exception:
        return None
    return FOREVER_GREEN_HEX if v > 0 else FOREVER_RED_HEX


def _hex_to_rgb01(hex_str: str) -> tuple[float, float, float]:
    """'1B4F91' → (0.106, 0.310, 0.569). Usato da reportlab."""
    h = hex_str.lstrip("#")
    return tuple(int(h[i:i + 2], 16) / 255.0 for i in (0, 2, 4))


# ─────────────────────────────────────────────────────────────────
#  CARICAMENTO MULTI-MESE
# ─────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_all_months(
    file_bytes: bytes,
    ambassador: str,
    regioni_csv: str,
    sheet_name: str,
    filter_mode: str,
) -> pd.DataFrame:
    """
    Carica TUTTI i mesi del file applicando il filtro zona, ma SENZA
    il filtro per mese latest. Restituisce un DataFrame con colonne MONTH/YEAR
    e tutte le metriche disponibili.

    Compatibile con la stessa struttura di filter_mode di load_and_filter
    in app.py (AND / OR / AMB / REG).
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)

    # Colonne minime richieste
    needed = ["SHOP_CODE", "DISTRICT_MANAGER", "MONTH", "YEAR", "AMBASSADOR", "REGION"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(
            f"Il foglio '{sheet_name}' non contiene queste colonne: {', '.join(missing)}"
        )

    # Cast numerici per ogni colonna metrica nota
    numeric_cols = (
        ["TAM", "SR_PLUS_GROSS_SALES", "SR_PLUS_NET_SALES",
         "SR_PLUS_GROSS_AR", "SR_PLUS_NET_AR",
         "SR_PLUS_OPTOUT", "SR_PLUS_OPTOUT_RATE",
         "R4_BASIC_GROSS_SALES", "R4_BASIC_NET_SALES",
         "R4_PREMIUM_GROSS_SALES", "R4_PREMIUM_NET_SALES",
         "EXCHANGE_SALES", "Activeforever", "TotalStore"]
        + [f["tam_col"] for f in FASCE_DEF]
        + [f["gross_col"] for f in FASCE_DEF]
    )
    for c in numeric_cols:
        if c not in df.columns:
            df[c] = 0
        else:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Cast string sicuri
    for c in ["COMPANY_NAME", "CITY", "STORE_TYPE", "STORE", "PROVINCE_CODE",
              "AREA_MANAGER", "DISTRICT_MANAGER", "AMBASSADOR", "REGION"]:
        if c not in df.columns:
            df[c] = ""
        else:
            df[c] = df[c].astype(object).where(df[c].notna(), "")

    valid = df[df["SHOP_CODE"].notna()].copy()

    amb_filter = [ambassador] if ambassador else []
    reg_filter = [r.strip() for r in regioni_csv.split(",") if r.strip()]

    if filter_mode == "AND":
        m = pd.Series(True, index=valid.index)
        if amb_filter:
            m &= valid["AMBASSADOR"].isin(amb_filter)
        if reg_filter:
            m &= valid["REGION"].isin(reg_filter)
    elif filter_mode == "OR":
        m = pd.Series(False, index=valid.index)
        if amb_filter:
            m |= valid["AMBASSADOR"].isin(amb_filter)
        if reg_filter:
            m |= valid["REGION"].isin(reg_filter)
        if not amb_filter and not reg_filter:
            m = pd.Series(True, index=valid.index)
    elif filter_mode == "AMB":
        m = valid["AMBASSADOR"].isin(amb_filter) if amb_filter else pd.Series(True, index=valid.index)
    else:  # REG
        m = valid["REGION"].isin(reg_filter) if reg_filter else pd.Series(True, index=valid.index)

    out = valid[m].copy()
    if out.empty:
        raise ValueError(
            "Nessun dato per la zona selezionata. Controlla Ambassador/Regioni/Modalità filtro."
        )

    out["MONTH"] = pd.to_numeric(out["MONTH"], errors="coerce").fillna(0).astype(int)
    out["YEAR"] = pd.to_numeric(out["YEAR"], errors="coerce").fillna(0).astype(int)
    return out


# ─────────────────────────────────────────────────────────────────
#  ESTRAZIONE METRICHE PER NEGOZIO/MESE
# ─────────────────────────────────────────────────────────────────
def _extract_metrics_row(row: pd.Series) -> dict[str, float]:
    """
    Estrae tutte le metriche calcolabili per un singolo record store/mese.
    Restituisce un dict con valori normalizzati.
    """
    tam = _safe_float(row.get("TAM"))
    gross = _safe_float(row.get("SR_PLUS_GROSS_SALES"))
    net = _safe_float(row.get("SR_PLUS_NET_SALES"))
    optout = _safe_float(row.get("SR_PLUS_OPTOUT"))

    metrics = {
        "qta": tam,
        "gross": gross,
        "ar_gross": _safe_pct(gross, tam),
        "net": net,
        "ar_net": _safe_pct(net, tam),
        "optout": optout,
        # Optout% = SR_PLUS_OPTOUT / TAM (se TAM>0). Se la colonna pre-calcolata
        # esiste e ha senso (>0), preferisco il calcolato perché coerente.
        "optout_pct": _safe_pct(optout, tam),
        "fb_gross": _safe_float(row.get("R4_BASIC_GROSS_SALES")),
        "fb_net": _safe_float(row.get("R4_BASIC_NET_SALES")),
        "fp_gross": _safe_float(row.get("R4_PREMIUM_GROSS_SALES")),
        "fp_net": _safe_float(row.get("R4_PREMIUM_NET_SALES")),
        "exchange": _safe_float(row.get("EXCHANGE_SALES")),
        "forever_active": _safe_float(row.get("Activeforever")),
    }
    # Fasce
    for f in FASCE_DEF:
        metrics[f"fascia_{f['key']}_tam"] = _safe_float(row.get(f["tam_col"]))
        metrics[f"fascia_{f['key']}_gross"] = _safe_float(row.get(f["gross_col"]))
    return metrics


def _aggregate_metrics(rows: pd.DataFrame) -> dict[str, float]:
    """Aggrega le metriche su un set di righe (per riga totale SA o store con più mesi)."""
    if rows.empty:
        return {k: 0.0 for k in [
            "qta", "gross", "ar_gross", "net", "ar_net", "optout", "optout_pct",
            "fb_gross", "fb_net", "fp_gross", "fp_net", "exchange", "forever_active",
        ]}
    tam = float(rows["TAM"].sum())
    gross = float(rows["SR_PLUS_GROSS_SALES"].sum())
    net = float(rows["SR_PLUS_NET_SALES"].sum())
    optout = float(rows["SR_PLUS_OPTOUT"].sum()) if "SR_PLUS_OPTOUT" in rows.columns else 0.0
    out = {
        "qta": tam,
        "gross": gross,
        "ar_gross": _safe_pct(gross, tam),
        "net": net,
        "ar_net": _safe_pct(net, tam),
        "optout": optout,
        "optout_pct": _safe_pct(optout, tam),
        "fb_gross": float(rows.get("R4_BASIC_GROSS_SALES", pd.Series([0])).sum()),
        "fb_net": float(rows.get("R4_BASIC_NET_SALES", pd.Series([0])).sum()),
        "fp_gross": float(rows.get("R4_PREMIUM_GROSS_SALES", pd.Series([0])).sum()),
        "fp_net": float(rows.get("R4_PREMIUM_NET_SALES", pd.Series([0])).sum()),
        "exchange": float(rows.get("EXCHANGE_SALES", pd.Series([0])).sum()),
        "forever_active": float(rows.get("Activeforever", pd.Series([0])).sum()),
    }
    for f in FASCE_DEF:
        out[f"fascia_{f['key']}_tam"] = float(rows.get(f["tam_col"], pd.Series([0])).sum())
        out[f"fascia_{f['key']}_gross"] = float(rows.get(f["gross_col"], pd.Series([0])).sum())
    return out


# ─────────────────────────────────────────────────────────────────
#  FORMATTAZIONE VALORI
# ─────────────────────────────────────────────────────────────────
def _fmt_metric(key: str, val: Any) -> str:
    """Formatta un valore in base al tipo di metrica."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    if key in ("ar_gross", "ar_net", "optout_pct"):
        return f"{val:.1f}%"
    # Quantità e numerici interi
    return f"{int(round(float(val))):,}".replace(",", ".")


def _fmt_eur(val: float) -> str:
    return f"€ {val:,.0f}".replace(",", ".")


# ─────────────────────────────────────────────────────────────────
#  COSTRUZIONE DATI PER UN SINGOLO SA
# ─────────────────────────────────────────────────────────────────
def build_sa_dataset(
    df_all: pd.DataFrame,
    sa_name: str,
    months: list[tuple[int, int]],
    columns_cfg: dict,
    fasce_cfg: dict,
    compensi_on: bool,
) -> dict:
    """
    Costruisce la struttura dati pronta per export per un singolo SA.

    Returns dict con:
      - sa_name
      - rows: list di dict con anagrafica + metriche per mese
      - totals: dict con totali aggregati per mese
      - months: list di (m, y)
      - columns_cfg, fasce_cfg, compensi_on (echo)
    """
    sub = df_all[df_all["DISTRICT_MANAGER"] == sa_name].copy()
    if sub.empty:
        return {
            "sa_name": sa_name, "rows": [], "totals": {}, "months": months,
            "columns_cfg": columns_cfg, "fasce_cfg": fasce_cfg,
            "compensi_on": compensi_on,
        }

    # Filtra solo i mesi richiesti
    if months:
        mask = pd.Series(False, index=sub.index)
        for (m, y) in months:
            mask |= (sub["MONTH"] == m) & (sub["YEAR"] == y)
        sub = sub[mask]
    if sub.empty:
        return {
            "sa_name": sa_name, "rows": [], "totals": {}, "months": months,
            "columns_cfg": columns_cfg, "fasce_cfg": fasce_cfg,
            "compensi_on": compensi_on,
        }

    # Lista univoca di store del SA con anagrafica (prendo la prima occorrenza)
    anagrafica_cols = ["SHOP_CODE", "STORE", "COMPANY_NAME", "CITY",
                       "PROVINCE_CODE", "STORE_TYPE", "REGION"]
    for c in anagrafica_cols:
        if c not in sub.columns:
            sub[c] = ""

    stores = sub.drop_duplicates("SHOP_CODE")[anagrafica_cols].copy()

    # Per ordinamento: Canale → Regione → NET% del mese più recente DESC
    # (NET% = SR_PLUS_NET_SALES / TAM aggregato sui mesi selezionati)
    last_month = max(months, key=lambda x: (x[1], x[0])) if months else None

    rows_data: list[dict] = []
    for _, sr in stores.iterrows():
        shop = sr["SHOP_CODE"]
        store_records = sub[sub["SHOP_CODE"] == shop]
        store_type = _safe_str(sr["STORE_TYPE"])
        canale = _channel_for_store_type(store_type)

        per_month: dict[tuple[int, int], dict] = {}
        for (m, y) in months:
            rec = store_records[(store_records["MONTH"] == m) & (store_records["YEAR"] == y)]
            if rec.empty:
                per_month[(m, y)] = {
                    k: 0.0 for k in [
                        "qta", "gross", "ar_gross", "net", "ar_net", "optout",
                        "optout_pct", "fb_gross", "fb_net", "fp_gross", "fp_net",
                        "exchange", "forever_active",
                    ]
                }
                for f in FASCE_DEF:
                    per_month[(m, y)][f"fascia_{f['key']}_tam"] = 0.0
                    per_month[(m, y)][f"fascia_{f['key']}_gross"] = 0.0
            else:
                per_month[(m, y)] = _extract_metrics_row(rec.iloc[0])

        # Per ordinamento globale uso il mese più recente
        sort_ar_net = per_month.get(last_month, {}).get("ar_net", 0.0) if last_month else 0.0

        # Compensi (calcolati sul mese più recente selezionato)
        comp_data = None
        if compensi_on and last_month is not None:
            metrics_last = per_month[last_month]
            net_last = metrics_last["net"]
            ar_net_last = metrics_last["ar_net"]
            mult = _multiplier_from_ar(ar_net_last, canale)
            base = net_last * COMPENSO_PER_PEZZO
            comp_data = {
                "canale_compenso": canale,
                "ar_net_riferimento": ar_net_last,
                "mese_riferimento": _month_label(*last_month),
                "tier_attuale": mult,
                "compenso_attuale": base * mult,
                "compenso_x2": base * 2,
                "compenso_x3": base * 3,
                "compenso_x4": base * 4,
            }

        rows_data.append({
            "shop_code": _safe_str(shop),
            "store": _safe_str(sr["STORE"]),
            "company": _safe_str(sr["COMPANY_NAME"]),
            "city": _safe_str(sr["CITY"]),
            "province": _safe_str(sr["PROVINCE_CODE"]),
            "store_type": store_type,
            "canale": canale,
            "region": _safe_str(sr["REGION"]),
            "metrics_by_month": per_month,
            "compensi": comp_data,
            "_sort_canale": canale,
            "_sort_region": _safe_str(sr["REGION"]),
            "_sort_ar_net": sort_ar_net,
        })

    # Sort: Canale → Regione → NET% desc
    canale_order = {"Franchising": 0, "Dealer": 1}
    rows_data.sort(key=lambda r: (
        canale_order.get(r["_sort_canale"], 99),
        r["_sort_region"],
        -r["_sort_ar_net"],
    ))

    # Totali per mese
    totals_by_month: dict[tuple[int, int], dict] = {}
    for (m, y) in months:
        rec = sub[(sub["MONTH"] == m) & (sub["YEAR"] == y)]
        totals_by_month[(m, y)] = _aggregate_metrics(rec)

    # Totale compensi (solo se attivi)
    totals_compensi = None
    if compensi_on:
        totals_compensi = {
            "compenso_attuale": sum((r.get("compensi") or {}).get("compenso_attuale", 0.0) for r in rows_data),
            "compenso_x2":      sum((r.get("compensi") or {}).get("compenso_x2",      0.0) for r in rows_data),
            "compenso_x3":      sum((r.get("compensi") or {}).get("compenso_x3",      0.0) for r in rows_data),
            "compenso_x4":      sum((r.get("compensi") or {}).get("compenso_x4",      0.0) for r in rows_data),
        }

    return {
        "sa_name": sa_name,
        "rows": rows_data,
        "totals_by_month": totals_by_month,
        "totals_compensi": totals_compensi,
        "months": months,
        "columns_cfg": columns_cfg,
        "fasce_cfg": fasce_cfg,
        "compensi_on": compensi_on,
    }


# ─────────────────────────────────────────────────────────────────
#  COSTRUZIONE LISTA DEFINITIVA DELLE COLONNE (per export e preview)
# ─────────────────────────────────────────────────────────────────
def build_column_plan(
    months: list[tuple[int, int]],
    columns_cfg: dict,
    fasce_cfg: dict,
    compensi_on: bool,
) -> list[dict]:
    """
    Restituisce la lista ordinata dei descrittori di colonna che andranno
    nell'output. Ogni descrittore è:
        {
          "id": str univoco,
          "label": str visibile,
          "group": "anagrafica" | "metrica" | "fascia" | "compenso",
          "month": (m, y) o None,
          "metric_key": str o None,
          "fascia_key": str o None,
          "fascia_part": "tam" | "gross" | None,
          "highlight_net": bool,
          "color_rule": "ar" | "optout" | "forever" | None,
        }
    """
    plan: list[dict] = []

    # Anagrafica fissa
    plan.append({"id": "shop_code", "label": "Codice DW9", "group": "anagrafica"})
    plan.append({"id": "store",     "label": "Negozio",    "group": "anagrafica"})
    plan.append({"id": "company",   "label": "Ragione Sociale", "group": "anagrafica"})
    plan.append({"id": "city",      "label": "Città",      "group": "anagrafica"})
    plan.append({"id": "store_type","label": "Tipo Store", "group": "anagrafica"})
    plan.append({"id": "canale",    "label": "Canale",     "group": "anagrafica"})
    plan.append({"id": "region",    "label": "Regione",    "group": "anagrafica"})

    # Metriche per mese
    selected_metrics = [m for m in METRICHE_BASE if columns_cfg.get(m["key"], False)]
    for (mi, yi) in months:
        for met in selected_metrics:
            color_rule = None
            if met["key"] in ("ar_gross", "ar_net"):
                color_rule = "ar"
            elif met["key"] == "optout_pct":
                color_rule = "optout"
            elif met["key"] in ("fb_net", "fp_net"):
                # forever-related: il colore "verde/rosso" si applica più
                # al numero di Forever attivi che alle vendite. Lo lascio neutro.
                color_rule = None
            plan.append({
                "id":           f"m_{met['key']}_{mi}_{yi}",
                "label":        met["label"],
                "group":        "metrica",
                "month":        (mi, yi),
                "metric_key":   met["key"],
                "color_rule":   color_rule,
                "highlight_net": met.get("highlight", False),
            })

        # Fasce per mese (TAM + Gross per ogni fascia selezionata)
        for f in FASCE_DEF:
            if not fasce_cfg.get(f["key"], False):
                continue
            plan.append({
                "id":         f"f_{f['key']}_tam_{mi}_{yi}",
                "label":      f"Q. {f['label']}",
                "group":      "fascia",
                "month":      (mi, yi),
                "fascia_key": f["key"],
                "fascia_part":"tam",
            })
            plan.append({
                "id":         f"f_{f['key']}_gross_{mi}_{yi}",
                "label":      f"Gross {f['label']}",
                "group":      "fascia",
                "month":      (mi, yi),
                "fascia_key": f["key"],
                "fascia_part":"gross",
            })

    # Compensi (sezione finale, calcolati sul mese più recente)
    if compensi_on:
        plan.append({"id": "comp_canale",   "label": "Canale (compenso)", "group": "compenso"})
        plan.append({"id": "comp_arnet",    "label": "AR Net% rif.",      "group": "compenso", "color_rule": "ar"})
        plan.append({"id": "comp_tier",     "label": "Tier attuale",      "group": "compenso"})
        plan.append({"id": "comp_attuale",  "label": "Compenso attuale (€)",  "group": "compenso", "compenso_match": "attuale"})
        plan.append({"id": "comp_x2",       "label": "Compenso X2 (€)",       "group": "compenso", "compenso_match": "x2"})
        plan.append({"id": "comp_x3",       "label": "Compenso X3 (€)",       "group": "compenso", "compenso_match": "x3"})
        plan.append({"id": "comp_x4",       "label": "Compenso X4 (€)",       "group": "compenso", "compenso_match": "x4"})

    return plan


def _value_for_column(row: dict, col: dict) -> tuple[Any, str]:
    """
    Estrae (valore_raw, valore_formattato) per un dato (riga, colonna).
    """
    g = col["group"]
    if g == "anagrafica":
        v = row.get(col["id"], "")
        return v, _safe_str(v)
    if g == "metrica":
        m_key = col["metric_key"]
        mv = row["metrics_by_month"].get(col["month"], {})
        raw = mv.get(m_key, 0.0)
        return raw, _fmt_metric(m_key, raw)
    if g == "fascia":
        mv = row["metrics_by_month"].get(col["month"], {})
        key = f"fascia_{col['fascia_key']}_{col['fascia_part']}"
        raw = mv.get(key, 0.0)
        return raw, _fmt_metric("qta", raw)
    if g == "compenso":
        comp = row.get("compensi") or {}
        cid = col["id"]
        if cid == "comp_canale":
            return comp.get("canale_compenso", ""), _safe_str(comp.get("canale_compenso", ""))
        if cid == "comp_arnet":
            v = comp.get("ar_net_riferimento", 0.0)
            return v, f"{v:.1f}%"
        if cid == "comp_tier":
            t = comp.get("tier_attuale", 1)
            return t, f"X{t}"
        if cid in ("comp_attuale", "comp_x2", "comp_x3", "comp_x4"):
            mp = {"comp_attuale": "compenso_attuale",
                  "comp_x2": "compenso_x2",
                  "comp_x3": "compenso_x3",
                  "comp_x4": "compenso_x4"}
            v = comp.get(mp[cid], 0.0)
            return v, _fmt_eur(v)
    return "", ""


def _value_for_total(totals_by_month: dict, totals_compensi: dict | None, col: dict) -> tuple[Any, str]:
    """Valore della riga totale per una data colonna."""
    g = col["group"]
    if g == "anagrafica":
        if col["id"] == "shop_code":
            return "TOTALE", "TOTALE"
        return "", ""
    if g == "metrica":
        mv = totals_by_month.get(col["month"], {})
        raw = mv.get(col["metric_key"], 0.0)
        return raw, _fmt_metric(col["metric_key"], raw)
    if g == "fascia":
        mv = totals_by_month.get(col["month"], {})
        key = f"fascia_{col['fascia_key']}_{col['fascia_part']}"
        raw = mv.get(key, 0.0)
        return raw, _fmt_metric("qta", raw)
    if g == "compenso":
        if totals_compensi is None:
            return "", ""
        cid = col["id"]
        if cid == "comp_canale":
            return "", "—"
        if cid == "comp_arnet":
            return "", "—"
        if cid == "comp_tier":
            return "", "—"
        if cid in ("comp_attuale", "comp_x2", "comp_x3", "comp_x4"):
            mp = {"comp_attuale": "compenso_attuale",
                  "comp_x2": "compenso_x2",
                  "comp_x3": "compenso_x3",
                  "comp_x4": "compenso_x4"}
            v = totals_compensi.get(mp[cid], 0.0)
            return v, _fmt_eur(v)
    return "", ""


# ─────────────────────────────────────────────────────────────────
#  DATAFRAME PER ANTEPRIMA
# ─────────────────────────────────────────────────────────────────
def build_preview_dataframe(sa_data: dict) -> pd.DataFrame:
    """Costruisce un DataFrame "piatto" pronto per st.dataframe."""
    plan = build_column_plan(
        sa_data["months"], sa_data["columns_cfg"],
        sa_data["fasce_cfg"], sa_data["compensi_on"],
    )

    headers = []
    for c in plan:
        if c["group"] == "metrica" or c["group"] == "fascia":
            mlabel = _month_label(*c["month"])
            headers.append(f"{mlabel} · {c['label']}")
        else:
            headers.append(c["label"])

    out_rows = []
    for r in sa_data["rows"]:
        row_vals = []
        for c in plan:
            _, fmt = _value_for_column(r, c)
            row_vals.append(fmt)
        out_rows.append(row_vals)

    # Riga totale
    if sa_data["rows"]:
        total_row = []
        for c in plan:
            _, fmt = _value_for_total(
                sa_data["totals_by_month"], sa_data["totals_compensi"], c
            )
            total_row.append(fmt)
        out_rows.append(total_row)

    return pd.DataFrame(out_rows, columns=headers)


# ─────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────
def build_excel_chiusura(sa_data: dict, period_label: str) -> io.BytesIO:
    """Costruisce un file Excel per un singolo SA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chiusura"

    plan = build_column_plan(
        sa_data["months"], sa_data["columns_cfg"],
        sa_data["fasce_cfg"], sa_data["compensi_on"],
    )

    # ── Stili ──
    HEADER_FILL = PatternFill("solid", fgColor=HEADER_HEX)
    NET_FILL = PatternFill("solid", fgColor=NET_HEADER_HEX)
    SUBHEADER_FILL = PatternFill("solid", fgColor=SUBHEADER_HEX)
    TOTAL_FILL = PatternFill("solid", fgColor=TOTAL_ROW_HEX)
    TIER_FILL = PatternFill("solid", fgColor=TIER_REACHED_HEX)
    BD = Border(left=Side(style="thin", color="888888"),
                right=Side(style="thin", color="888888"),
                top=Side(style="thin", color="888888"),
                bottom=Side(style="thin", color="888888"))
    BD_THICK = Border(left=Side(style="medium", color="000000"),
                      right=Side(style="medium", color="000000"),
                      top=Side(style="thin", color="888888"),
                      bottom=Side(style="thin", color="888888"))
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    H_FONT = Font(bold=True, color="FFFFFF", size=10)
    H_FONT_NET = Font(bold=True, color="FFFFFF", size=10)
    H_FONT_GROUP = Font(bold=True, color="FFFFFF", size=11)
    REG_FONT = Font(size=10)
    TOT_FONT = Font(bold=True, size=10, color="003087")
    NET_BODY_FONT = Font(bold=True, size=10)

    # ── Titolo ──
    n_cols = len(plan)
    last_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"CHIUSURA — {sa_data['sa_name']} · {period_label}"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = C
    ws.row_dimensions[1].height = 32

    # ── Riga 2: GROUP HEADER (mesi + sezioni) ──
    # Costruisco gruppi consecutivi per mese / sezione
    groups: list[tuple[str, int, int, str | None]] = []  # (label, col_start, col_end, hex_color)
    i = 0
    months_seen = []
    while i < len(plan):
        c = plan[i]
        if c["group"] == "anagrafica":
            j = i
            while j < len(plan) and plan[j]["group"] == "anagrafica":
                j += 1
            groups.append(("Anagrafica", i + 1, j, HEADER_HEX))
            i = j
        elif c["group"] in ("metrica", "fascia"):
            month = c["month"]
            j = i
            while j < len(plan) and plan[j]["group"] in ("metrica", "fascia") and plan[j]["month"] == month:
                j += 1
            month_idx = months_seen.index(month) if month in months_seen else len(months_seen)
            if month not in months_seen:
                months_seen.append(month)
            color = MONTH_COLORS_HEX[month_idx % len(MONTH_COLORS_HEX)]
            groups.append((_month_label(*month).upper(), i + 1, j, color))
            i = j
        elif c["group"] == "compenso":
            j = i
            while j < len(plan) and plan[j]["group"] == "compenso":
                j += 1
            groups.append(("COMPENSI (mese di rif.)", i + 1, j, "8B5A2B"))
            i = j
        else:
            i += 1

    for (label, c1, c2, hex_color) in groups:
        if c1 == c2:
            ws.cell(row=2, column=c1, value=label)
            cell = ws.cell(row=2, column=c1)
        else:
            ws.merge_cells(start_row=2, end_row=2, start_column=c1, end_column=c2)
            cell = ws.cell(row=2, column=c1, value=label)
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.font = H_FONT_GROUP
        cell.alignment = C
        cell.border = BD_THICK
    ws.row_dimensions[2].height = 24

    # ── Riga 3: LABEL COLONNA ──
    for ci, col in enumerate(plan, 1):
        cell = ws.cell(row=3, column=ci, value=col["label"])
        cell.alignment = C
        cell.border = BD_THICK if (col["group"] in ("metrica", "fascia") and ci > 1
                                   and plan[ci - 2]["group"] != col["group"]) else BD

        # NET% e Net Sales: header verde scuro
        is_net_header = (col.get("highlight_net") is True) or \
                        (col["group"] == "metrica" and col["metric_key"] == "net")
        if is_net_header:
            cell.fill = NET_FILL
            cell.font = H_FONT_NET
        elif col["group"] == "compenso":
            cell.fill = SUBHEADER_FILL
            cell.font = H_FONT
        elif col["group"] == "anagrafica":
            cell.fill = HEADER_FILL
            cell.font = H_FONT
        else:
            cell.fill = SUBHEADER_FILL
            cell.font = H_FONT
    ws.row_dimensions[3].height = 30

    # ── Righe dati ──
    base_row = 4
    for ri, r in enumerate(sa_data["rows"]):
        excel_row = base_row + ri
        for ci, col in enumerate(plan, 1):
            raw, fmt = _value_for_column(r, col)
            cell = ws.cell(row=excel_row, column=ci, value=fmt)
            cell.border = BD
            cell.alignment = C if col["group"] != "anagrafica" else L

            # Bold per NET e NET%
            if (col.get("highlight_net") is True) or \
               (col["group"] == "metrica" and col["metric_key"] == "net"):
                cell.font = NET_BODY_FONT
            else:
                cell.font = REG_FONT

            # Semaforo
            rule = col.get("color_rule")
            if rule == "ar" and isinstance(raw, (int, float)):
                hexc = _color_ar_hex(raw)
                if hexc:
                    cell.fill = PatternFill("solid", fgColor=hexc)
            elif rule == "optout" and isinstance(raw, (int, float)):
                hexc = _color_optout_hex(raw)
                if hexc:
                    cell.fill = PatternFill("solid", fgColor=hexc)

            # Forever cell: highlight binario (sulle Forever attive aggregate
            # del mese — calcolato a parte dato che non è una colonna esposta).
            # Decisione: applichiamo l'highlight su fb_net + fp_net se attivati.

            # Compenso tier raggiunto: evidenzia in verde la colonna corrispondente
            if col["group"] == "compenso":
                comp = r.get("compensi") or {}
                tier = comp.get("tier_attuale", 0)
                match = col.get("compenso_match")
                if match:
                    expected_tier = {"attuale": tier, "x2": 2, "x3": 3, "x4": 4}.get(match)
                    if expected_tier == tier and expected_tier != 0:
                        cell.fill = TIER_FILL
                        cell.font = Font(bold=True, size=10)

        ws.row_dimensions[excel_row].height = 20

    # ── Riga totali SA ──
    if sa_data["rows"]:
        tot_row = base_row + len(sa_data["rows"])
        for ci, col in enumerate(plan, 1):
            raw, fmt = _value_for_total(
                sa_data["totals_by_month"], sa_data["totals_compensi"], col
            )
            cell = ws.cell(row=tot_row, column=ci, value=fmt)
            cell.border = BD_THICK
            cell.fill = TOTAL_FILL
            cell.alignment = C if col["group"] != "anagrafica" else L

            # Bold NET nei totali
            if (col.get("highlight_net") is True) or \
               (col["group"] == "metrica" and col["metric_key"] == "net"):
                cell.font = Font(bold=True, size=11, color="003087")
            else:
                cell.font = TOT_FONT

            # Semaforo anche sui totali (per AR%/Optout%)
            rule = col.get("color_rule")
            if rule == "ar" and isinstance(raw, (int, float)):
                hexc = _color_ar_hex(raw)
                if hexc:
                    cell.fill = PatternFill("solid", fgColor=hexc)
            elif rule == "optout" and isinstance(raw, (int, float)):
                hexc = _color_optout_hex(raw)
                if hexc:
                    cell.fill = PatternFill("solid", fgColor=hexc)
        ws.row_dimensions[tot_row].height = 24

    # ── Larghezza colonne automatica ──
    for col_cells in ws.columns:
        try:
            letter = get_column_letter(col_cells[0].column)
        except Exception:
            continue
        max_len = 0
        for c in col_cells:
            try:
                v = c.value
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            except Exception:
                continue
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 28)

    # Freeze panes su righe header e prima colonna (DW9)
    ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
#  PDF EXPORT (reportlab, landscape A4)
# ─────────────────────────────────────────────────────────────────
def build_pdf_chiusura(sa_data: dict, period_label: str) -> bytes:
    """Costruisce un PDF landscape A4 per un singolo SA."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Chiusura {sa_data['sa_name']}",
    )
    styles = getSampleStyleSheet()
    story = []

    # Titolo
    title_style = styles["Title"]
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor(f"#{HEADER_HEX}")
    story.append(Paragraph(
        f"<b>CHIUSURA — {sa_data['sa_name']}</b>",
        title_style,
    ))
    sub_style = styles["Normal"]
    sub_style.fontSize = 9
    sub_style.textColor = colors.grey
    story.append(Paragraph(period_label, sub_style))
    story.append(Spacer(1, 4 * mm))

    if not sa_data["rows"]:
        story.append(Paragraph("Nessun dato disponibile per i mesi selezionati.", styles["Normal"]))
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    plan = build_column_plan(
        sa_data["months"], sa_data["columns_cfg"],
        sa_data["fasce_cfg"], sa_data["compensi_on"],
    )

    # ── Header a 2 livelli ──
    # Riga 1: gruppi (anagrafica, mese, compensi)
    # Riga 2: label colonna
    months_seen = []
    group_row = []
    label_row = []
    group_spans = []  # list di (label, c1, c2, hex_color)
    i = 0
    while i < len(plan):
        c = plan[i]
        if c["group"] == "anagrafica":
            j = i
            while j < len(plan) and plan[j]["group"] == "anagrafica":
                j += 1
            group_spans.append(("Anagrafica", i, j - 1, HEADER_HEX))
            i = j
        elif c["group"] in ("metrica", "fascia"):
            month = c["month"]
            j = i
            while j < len(plan) and plan[j]["group"] in ("metrica", "fascia") and plan[j]["month"] == month:
                j += 1
            if month not in months_seen:
                months_seen.append(month)
            color = MONTH_COLORS_HEX[(months_seen.index(month)) % len(MONTH_COLORS_HEX)]
            group_spans.append((_month_label(*month).upper(), i, j - 1, color))
            i = j
        else:  # compenso
            j = i
            while j < len(plan) and plan[j]["group"] == "compenso":
                j += 1
            group_spans.append(("COMPENSI", i, j - 1, "8B5A2B"))
            i = j

    n_cols = len(plan)
    group_row = [""] * n_cols
    for (label, c1, c2, hexc) in group_spans:
        group_row[c1] = label

    # Stile Paragraph per header colonna (word-wrap automatico, white bold)
    header_para_style = styles["Normal"].clone("header_para")
    header_para_style.fontSize = 6.8
    header_para_style.leading = 7.8
    header_para_style.alignment = 1  # CENTER
    header_para_style.textColor = colors.white
    header_para_style.fontName = "Helvetica-Bold"

    def _wrap_header(label: str) -> Paragraph:
        return Paragraph(str(label).replace("&", "&amp;").replace("€", "EUR"), header_para_style)

    label_row = [_wrap_header(c["label"]) for c in plan]

    # ── Righe dati ──
    # Stile Paragraph compatto per le celle anagrafiche (word-wrap automatico)
    cell_para_style = styles["Normal"].clone("cell_para")
    cell_para_style.fontSize = 6.5
    cell_para_style.leading = 7.5
    cell_para_style.alignment = 1  # CENTER
    cell_para_style.textColor = colors.black

    def _wrap_anagrafica(val: str) -> Paragraph:
        """Avvolge il testo in un Paragraph per word-wrap automatico."""
        return Paragraph(str(val).replace("&", "&amp;"), cell_para_style)

    data_rows = []
    for r in sa_data["rows"]:
        row_vals = []
        for c in plan:
            _, fmt = _value_for_column(r, c)
            # Usa Paragraph (con word-wrap) per le anagrafiche testuali lunghe
            if c["group"] == "anagrafica" and c["id"] in ("store", "company", "city"):
                row_vals.append(_wrap_anagrafica(fmt))
            else:
                row_vals.append(fmt)
        data_rows.append(row_vals)

    # Riga totali
    tot_vals = []
    for c in plan:
        _, fmt = _value_for_total(
            sa_data["totals_by_month"], sa_data["totals_compensi"], c
        )
        tot_vals.append(fmt)

    # Tabella reportlab
    table_data = [group_row, label_row] + data_rows + [tot_vals]

    # Calcolo larghezze colonne in base al contenuto dell'header
    # Larghezza disponibile: A4 landscape ~ 277mm - 20mm margini = 257mm
    avail_w = (landscape(A4)[0] - 20 * mm)
    # Larghezza minima per colonna in base al ruolo
    base_widths = []
    for c in plan:
        if c["group"] == "anagrafica":
            if c["id"] in ("shop_code",):
                base_widths.append(20)
            elif c["id"] in ("store", "company"):
                base_widths.append(28)
            elif c["id"] in ("city",):
                base_widths.append(20)
            elif c["id"] in ("region",):
                base_widths.append(18)
            elif c["id"] in ("store_type", "canale"):
                base_widths.append(18)
            else:
                base_widths.append(18)
        elif c["group"] == "metrica":
            base_widths.append(15 if c["metric_key"] in ("ar_gross", "ar_net", "optout_pct") else 14)
        elif c["group"] == "fascia":
            base_widths.append(13)
        else:
            base_widths.append(20)

    total_units = sum(base_widths)
    col_widths = [w / total_units * avail_w for w in base_widths]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=2)

    # ── Style table ──
    style_cmds = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 7.5),
        ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 2), (-1, -1), 6.8),
        ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#888888")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]

    # Group row spans + colors
    for (label, c1, c2, hexc) in group_spans:
        if c1 != c2:
            style_cmds.append(("SPAN", (c1, 0), (c2, 0)))
        style_cmds.append(("BACKGROUND", (c1, 0), (c2, 0), colors.HexColor(f"#{hexc}")))

    # Label row colors per colonna
    for ci, col in enumerate(plan):
        is_net_header = (col.get("highlight_net") is True) or \
                        (col["group"] == "metrica" and col["metric_key"] == "net")
        if is_net_header:
            style_cmds.append(("BACKGROUND", (ci, 1), (ci, 1), colors.HexColor(f"#{NET_HEADER_HEX}")))
        elif col["group"] == "anagrafica":
            style_cmds.append(("BACKGROUND", (ci, 1), (ci, 1), colors.HexColor(f"#{HEADER_HEX}")))
        else:
            style_cmds.append(("BACKGROUND", (ci, 1), (ci, 1), colors.HexColor(f"#{SUBHEADER_HEX}")))

    # Body: bold per Net / Net%, semaforo sulle %
    n_header_rows = 2
    for ri, r in enumerate(sa_data["rows"]):
        body_row = n_header_rows + ri
        for ci, col in enumerate(plan):
            raw, _ = _value_for_column(r, col)

            # Bold NET / NET%
            is_net = (col.get("highlight_net") is True) or \
                     (col["group"] == "metrica" and col["metric_key"] == "net")
            if is_net:
                style_cmds.append(("FONTNAME", (ci, body_row), (ci, body_row), "Helvetica-Bold"))

            # Semaforo
            rule = col.get("color_rule")
            if rule == "ar" and isinstance(raw, (int, float)):
                hexc = _color_ar_hex(raw)
                if hexc:
                    style_cmds.append(("BACKGROUND", (ci, body_row), (ci, body_row), colors.HexColor(f"#{hexc}")))
            elif rule == "optout" and isinstance(raw, (int, float)):
                hexc = _color_optout_hex(raw)
                if hexc:
                    style_cmds.append(("BACKGROUND", (ci, body_row), (ci, body_row), colors.HexColor(f"#{hexc}")))

            # Compenso tier
            if col["group"] == "compenso":
                comp = r.get("compensi") or {}
                tier = comp.get("tier_attuale", 0)
                match = col.get("compenso_match")
                if match:
                    expected = {"attuale": tier, "x2": 2, "x3": 3, "x4": 4}.get(match)
                    if expected == tier and expected != 0:
                        style_cmds.append(("BACKGROUND", (ci, body_row), (ci, body_row), colors.HexColor(f"#{TIER_REACHED_HEX}")))
                        style_cmds.append(("FONTNAME", (ci, body_row), (ci, body_row), "Helvetica-Bold"))

    # Riga totali (ultima)
    tot_idx = n_header_rows + len(sa_data["rows"])
    style_cmds.append(("BACKGROUND", (0, tot_idx), (-1, tot_idx), colors.HexColor(f"#{TOTAL_ROW_HEX}")))
    style_cmds.append(("FONTNAME", (0, tot_idx), (-1, tot_idx), "Helvetica-Bold"))
    style_cmds.append(("LINEABOVE", (0, tot_idx), (-1, tot_idx), 1.2, colors.HexColor(f"#{HEADER_HEX}")))

    # Bold NET nelle totali
    for ci, col in enumerate(plan):
        is_net = (col.get("highlight_net") is True) or \
                 (col["group"] == "metrica" and col["metric_key"] == "net")
        if is_net:
            style_cmds.append(("TEXTCOLOR", (ci, tot_idx), (ci, tot_idx), colors.HexColor(f"#{NET_HEADER_HEX}")))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} · Wind3 Reload Dashboard",
        sub_style,
    ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  STREAMLIT UI — RENDER TAB
# ─────────────────────────────────────────────────────────────────
def render_chiusure_tab(
    file_bytes: bytes,
    ambassador: str,
    regioni: list[str],
    sheet_name: str,
    filter_mode: str,
):
    """
    Render della pagina "Genera Chiusure". Va chiamata dentro un
    `with tab_x:` o come contenuto di una pagina nella sidebar.
    """
    st.markdown(
        '<div class="section-title">🧾 Genera chiusure mensili personalizzate</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Crea chiusure Excel/PDF specifiche per uno o più Sales Account, "
        "scegliendo metriche, mesi, fasce di prezzo e moltiplicatori compenso."
    )

    # Carica tutti i mesi (cache)
    try:
        df_all = load_all_months(
            file_bytes, ambassador, ",".join(regioni), sheet_name, filter_mode,
        )
    except Exception as e:
        st.error(f"Errore caricamento dati: {e}")
        return

    # Liste disponibili
    sa_list_all = sorted(df_all["DISTRICT_MANAGER"].dropna().unique().tolist())
    months_available = sorted(
        df_all[["MONTH", "YEAR"]]
        .drop_duplicates()
        .itertuples(index=False, name=None),
        key=lambda x: (x[1], x[0]),
    )

    if not sa_list_all:
        st.warning("Nessun Sales Account/DM trovato nei dati filtrati.")
        return
    if not months_available:
        st.warning("Nessun mese trovato nei dati.")
        return

    # ── 1. SELEZIONE SA ──
    st.markdown("### 1️⃣ Seleziona Sales Account")
    col_sa1, col_sa2 = st.columns([1, 2])
    with col_sa1:
        all_sa = st.checkbox("Tutti i SA", value=True, key="chiusure_all_sa")
    with col_sa2:
        if all_sa:
            sa_selected = sa_list_all
            st.info(f"Selezionati tutti i {len(sa_list_all)} SA")
        else:
            sa_selected = st.multiselect(
                "Scegli i SA",
                options=sa_list_all,
                default=sa_list_all[:3],
                key="chiusure_sa_select",
            )

    if not sa_selected:
        st.warning("Seleziona almeno un Sales Account per continuare.")
        return

    # ── 2. SELEZIONE MESI ──
    st.markdown("### 2️⃣ Seleziona mesi")
    n_cols_per_row = 4
    months_selected = []
    for batch_start in range(0, len(months_available), n_cols_per_row):
        cols = st.columns(n_cols_per_row)
        for i, (m, y) in enumerate(months_available[batch_start:batch_start + n_cols_per_row]):
            with cols[i]:
                # Default: ultimi 2 mesi
                default = (m, y) in months_available[-2:]
                if st.checkbox(
                    _month_label(m, y),
                    value=default,
                    key=f"chiusure_m_{m}_{y}",
                ):
                    months_selected.append((m, y))

    months_selected = sorted(months_selected, key=lambda x: (x[1], x[0]))

    if not months_selected:
        st.warning("Seleziona almeno un mese.")
        return

    # ── 3. CONFIGURAZIONE COLONNE ──
    st.markdown("### 3️⃣ Metriche da includere")
    st.caption("Scegli quali metriche far apparire per ogni mese selezionato.")

    columns_cfg = {}
    cols_metric = st.columns(3)
    for i, met in enumerate(METRICHE_BASE):
        with cols_metric[i % 3]:
            label = met["label"]
            if met.get("highlight"):
                label = f"⭐ **{label}**"
            columns_cfg[met["key"]] = st.checkbox(
                label,
                value=met["default"],
                key=f"chiusure_col_{met['key']}",
            )

    # ── 4. FASCE DI PREZZO ──
    st.markdown("### 4️⃣ Fasce di prezzo (TAM vs Gross per ogni fascia)")
    st.caption("Per ogni fascia attivata vengono aggiunte 2 colonne: Quantità e Gross.")
    fasce_cfg = {}
    cols_f = st.columns(3)
    for i, f in enumerate(FASCE_DEF):
        with cols_f[i % 3]:
            fasce_cfg[f["key"]] = st.checkbox(
                f["label"],
                value=False,
                key=f"chiusure_f_{f['key']}",
            )

    # ── 5. COMPENSI ──
    st.markdown("### 5️⃣ Moltiplicatori compenso")
    compensi_on = st.checkbox(
        "Includi sezione compensi (X1/X2/X3/X4 calcolata sul mese più recente selezionato)",
        value=False, key="chiusure_compensi",
    )
    if compensi_on:
        st.caption(
            "**Franchising:** X2 ≥ 40% AR, X3 ≥ 60%, X4 ≥ 75% · "
            "**Dealer / Large Chain / Owned:** X2 ≥ 25%, X3 ≥ 40%, X4 ≥ 55% · "
            "**Compenso = Net × 5€ × moltiplicatore**. "
            "La colonna del tier attualmente raggiunto è evidenziata in verde."
        )

    # ── 6. FORMATO OUTPUT ──
    st.markdown("### 6️⃣ Formato output")
    formato = st.radio(
        "Scegli formato",
        ["Excel (.xlsx)", "PDF (landscape A4)"],
        horizontal=True,
        key="chiusure_formato",
    )

    # ── 7. ANTEPRIMA ──
    st.markdown("---")
    st.markdown("### 👁️ Anteprima")

    # Verifica che almeno una metrica sia selezionata (oltre alle anagrafiche)
    n_metric_selected = sum(1 for v in columns_cfg.values() if v)
    n_fascia_selected = sum(1 for v in fasce_cfg.values() if v)
    if n_metric_selected == 0 and n_fascia_selected == 0 and not compensi_on:
        st.warning("Seleziona almeno una metrica, una fascia o attiva i compensi.")
        return

    # Pre-calcola i dataset di tutti i SA selezionati
    sa_datasets: list[dict] = []
    for sa_name in sa_selected:
        ds = build_sa_dataset(
            df_all, sa_name, months_selected, columns_cfg, fasce_cfg, compensi_on,
        )
        sa_datasets.append(ds)

    # Selectbox per scegliere il SA da preview
    sa_preview = st.selectbox(
        "SA da visualizzare in anteprima",
        options=[ds["sa_name"] for ds in sa_datasets],
        key="chiusure_preview_sa",
    )
    preview_ds = next((ds for ds in sa_datasets if ds["sa_name"] == sa_preview), None)
    if preview_ds is None or not preview_ds["rows"]:
        st.info("Nessun negozio per il SA selezionato nei mesi indicati.")
    else:
        preview_df = build_preview_dataframe(preview_ds)
        st.markdown(
            f"**{len(preview_ds['rows'])} negozi** · "
            f"{len(months_selected)} mese/i · "
            f"{len(preview_df.columns)} colonne totali"
        )
        st.dataframe(preview_df, use_container_width=True, height=400, hide_index=True)

    # ── 8. GENERAZIONE FILE ──
    st.markdown("---")
    st.markdown("### ⬇️ Download")

    period_label = " · ".join(_month_label(m, y) for (m, y) in months_selected)
    is_excel = formato.startswith("Excel")
    ext = "xlsx" if is_excel else "pdf"
    mime = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if is_excel else "application/pdf")

    col_d1, col_d2 = st.columns(2)

    # Download SA singolo (quello in preview)
    with col_d1:
        st.markdown("**File singolo SA**")
        if preview_ds is not None and preview_ds["rows"]:
            try:
                if is_excel:
                    file_data = build_excel_chiusura(preview_ds, period_label).getvalue()
                else:
                    file_data = build_pdf_chiusura(preview_ds, period_label)
                safe_name = (
                    preview_ds["sa_name"].replace("/", "_").replace("\\", "_")
                    .replace(" ", "_")[:50]
                )
                fname = f"chiusura_{safe_name}_{datetime.now().strftime('%Y%m%d')}.{ext}"
                st.download_button(
                    f"⬇️ Scarica {preview_ds['sa_name']} ({ext.upper()})",
                    data=file_data,
                    file_name=fname,
                    mime=mime,
                    use_container_width=True,
                    key=f"chiusure_dl_single_{ext}",
                )
            except Exception as e:
                st.error(f"Errore generazione file singolo: {e}")
        else:
            st.info("Nessun dato per il SA selezionato.")

    # Download ZIP
    with col_d2:
        st.markdown(f"**ZIP — tutti i {len(sa_datasets)} SA**")
        if st.button(
            f"🗜️ Genera ZIP ({ext.upper()})",
            use_container_width=True,
            key=f"chiusure_zip_btn_{ext}",
        ):
            with st.spinner("Generazione files..."):
                zip_buf = io.BytesIO()
                n_ok = 0
                n_err = 0
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for ds in sa_datasets:
                        if not ds["rows"]:
                            continue
                        try:
                            if is_excel:
                                fdata = build_excel_chiusura(ds, period_label).getvalue()
                            else:
                                fdata = build_pdf_chiusura(ds, period_label)
                            safe = (
                                ds["sa_name"].replace("/", "_").replace("\\", "_")
                                .replace(" ", "_")[:50]
                            )
                            fname = f"chiusura_{safe}.{ext}"
                            zf.writestr(fname, fdata)
                            n_ok += 1
                        except Exception as e:
                            n_err += 1
                            zf.writestr(
                                f"_ERRORE_{ds['sa_name'][:30]}.txt",
                                f"Errore generando il file: {e}",
                            )
                zip_buf.seek(0)
                st.success(f"✅ Generati {n_ok} file" + (f" · {n_err} errori" if n_err else ""))
                st.download_button(
                    "⬇️ Scarica ZIP completo",
                    data=zip_buf,
                    file_name=f"chiusure_{ambassador}_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key=f"chiusure_dl_zip_{ext}",
                )

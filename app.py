"""
Wind3 Reload Dashboard — v4
Autore: Salvo (Ambassador A6 — Campania & Puglia)

Novità v4:
- Bug fix completi (filtro AND, divisione per zero, soglie hard-coded, ecc.)
- Anagrafica gerarchica dinamica: Zona → Area Manager → District Manager → Store
- Filtri multi-livello: Zona / AM / DM / categoria store / singolo store
- Vista "Mappa Zona" con tutti i negozi che competono
- Soglie davvero configurabili (anche in Excel/PDF)
- Cache più solida, format_func ottimizzato
"""

import streamlit as st
import pandas as pd
import io
import zipfile
import json as _json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ─────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Wind3 Reload – Dashboard",
    page_icon="📶",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────────────────────────────
for k, v in {
    "selected_dm": None,
    "goto_detail": False,
    "goto_store_detail": False,
    "selected_store": None,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@300;400;500;600;700;800&family=Barlow+Condensed:wght@600;700;800&display=swap');
:root {
    --blue:#003087; --blue-mid:#0050B3; --blue-light:#1976D2;
    --accent:#00A8E0; --red:#C62828; --orange:#E65100;
    --bg:#04101F; --card:#081828; --border:rgba(0,168,224,0.18);
    --text:#E8F0FE; --muted:#7B97BF;
}
* { font-family:'Barlow',sans-serif !important; box-sizing:border-box; }
html,body,[data-testid="stApp"] { background:var(--bg) !important; color:var(--text) !important; }
#MainMenu,footer,header,.stDeployButton { display:none !important; }
[data-testid="stToolbar"] { display:none !important; }
.block-container { padding:1.2rem 2rem 3rem !important; max-width:1500px !important; }
[data-testid="stSidebar"] { background:var(--card) !important; border-right:1px solid var(--border) !important; }
[data-testid="stSidebar"] * { color:var(--text) !important; }
.w3-header {
    background:linear-gradient(135deg,#001850 0%,#003087 60%,#005CB8 100%);
    border-radius:14px; padding:1.6rem 2.2rem; margin-bottom:1.2rem;
    border:1px solid rgba(0,168,224,0.3); position:relative; overflow:hidden;
}
.w3-header::after {
    content:''; position:absolute; bottom:0; left:0; right:0; height:3px;
    background:linear-gradient(90deg,var(--accent),var(--blue-light),transparent);
}
.w3-glow { position:absolute; top:-60px; right:-60px; width:300px; height:300px;
    background:radial-gradient(circle,rgba(0,168,224,0.12) 0%,transparent 70%); pointer-events:none; }
.w3-title { font-family:'Barlow Condensed',sans-serif !important; font-size:2.1rem; font-weight:800; color:white; line-height:1; }
.w3-title span { color:var(--accent); }
.w3-sub { font-size:0.78rem; color:rgba(255,255,255,0.5); letter-spacing:0.14em; text-transform:uppercase; margin-top:0.35rem; }
.w3-badge { display:inline-block; background:rgba(0,168,224,0.15); border:1px solid var(--accent);
    color:var(--accent); font-size:0.72rem; font-weight:700; letter-spacing:0.12em;
    padding:0.2rem 0.65rem; border-radius:100px; text-transform:uppercase; margin-top:0.6rem; margin-right:0.4rem; }
.kpi-row { display:flex; gap:1rem; margin-bottom:1.2rem; flex-wrap:wrap; }
.kpi-tile { flex:1; min-width:140px; background:var(--card); border:1px solid var(--border);
    border-radius:12px; padding:1rem 1.3rem; position:relative; overflow:hidden; }
.kpi-tile::before { content:''; position:absolute; top:0; left:0; right:0; height:3px; }
.kpi-tile.accent::before { background:var(--accent); }
.kpi-tile.green::before  { background:#4CAF50; }
.kpi-tile.red::before    { background:var(--red); }
.kpi-tile.orange::before { background:var(--orange); }
.kpi-label { font-size:0.7rem; font-weight:700; letter-spacing:0.13em; text-transform:uppercase; color:var(--muted); margin-bottom:0.3rem; }
.kpi-value { font-family:'Barlow Condensed',sans-serif !important; font-size:1.9rem; font-weight:800; line-height:1; color:var(--text); }
.kpi-value.red   { color:#EF5350; }
.kpi-value.green { color:#66BB6A; }
.kpi-value.accent { color:var(--accent); }
.kpi-sub { font-size:0.72rem; color:var(--muted); margin-top:0.2rem; }
.section-title { font-family:'Barlow Condensed',sans-serif !important; font-size:0.72rem; font-weight:700;
    letter-spacing:0.16em; text-transform:uppercase; color:var(--accent);
    margin-bottom:0.75rem; padding-bottom:0.4rem; border-bottom:1px solid var(--border); margin-top:0.5rem; }
.dm-card { background:var(--card); border:1px solid var(--border); border-radius:10px;
    padding:0.85rem 1.2rem; display:flex; align-items:center; gap:1.2rem; margin-bottom:0.5rem;
    transition:border-color 0.2s; position:relative; overflow:hidden; }
.dm-card:hover { border-color:rgba(0,168,224,0.4); }
.dm-card.critico    { border-left:4px solid var(--red); }
.dm-card.attenzione { border-left:4px solid var(--orange); }
.dm-card.ok         { border-left:4px solid #4CAF50; }
.dm-name { font-weight:700; font-size:0.9rem; color:var(--text); min-width:220px; flex:1; }
.dm-region { font-size:0.72rem; color:var(--muted); font-weight:400; }
.dm-kpi { text-align:center; min-width:80px; }
.dm-kpi-label { font-size:0.65rem; color:var(--muted); letter-spacing:0.1em; text-transform:uppercase; }
.dm-kpi-val { font-family:'Barlow Condensed',sans-serif !important; font-size:1.3rem; font-weight:800; line-height:1.1; }
.dm-kpi-val.red    { color:#EF5350; }
.dm-kpi-val.orange { color:#FFA726; }
.dm-kpi-val.green  { color:#66BB6A; }
.dm-kpi-val.white  { color:var(--text); }
.dm-trend { font-size:0.7rem; }
.dm-trend.up   { color:#66BB6A; }
.dm-trend.down { color:#EF5350; }
.dm-trend.flat { color:var(--muted); }
.dm-alerts { font-size:0.72rem; color:#EF5350; min-width:160px; }
.dm-forever { font-size:0.7rem; color:var(--muted); text-align:center; min-width:80px; }
.tree-am { background:var(--card); border:1px solid var(--border); border-left:4px solid var(--accent);
    border-radius:10px; padding:0.8rem 1.2rem; margin-top:0.8rem; }
.tree-am-name { font-family:'Barlow Condensed',sans-serif !important; font-size:1.1rem; font-weight:800;
    color:var(--accent); letter-spacing:0.04em; }
.tree-dm { background:rgba(0,168,224,0.04); border-left:3px solid var(--blue-light);
    padding:0.5rem 1rem; margin:0.3rem 0 0.3rem 1.2rem; border-radius:6px; font-size:0.85rem; }
.tree-dm-name { font-weight:700; color:var(--text); }
.tree-dm-meta { font-size:0.72rem; color:var(--muted); margin-top:0.1rem; }
.tree-store { font-size:0.78rem; color:var(--muted); margin-left:2.5rem; padding:0.15rem 0; }
[data-testid="stFileUploader"] { background:var(--card) !important; border:2px dashed var(--border) !important; border-radius:12px !important; padding:1rem !important; }
.stDownloadButton > button { background:linear-gradient(135deg,var(--blue-mid),var(--blue-light)) !important;
    color:white !important; border:none !important; border-radius:8px !important; font-weight:600 !important; padding:0.5rem 1.2rem !important; }
.stDownloadButton > button:hover { background:linear-gradient(135deg,var(--blue-light),var(--accent)) !important; }
.stTabs [data-baseweb="tab-list"] { background:var(--card) !important; border-radius:10px !important; border:1px solid var(--border) !important; gap:0 !important; }
.stTabs [data-baseweb="tab"] { background:transparent !important; color:var(--muted) !important; font-weight:600 !important; font-size:0.85rem !important; border-radius:8px !important; }
.stTabs [aria-selected="true"] { background:var(--blue-mid) !important; color:white !important; }
.stTabs [data-baseweb="tab-panel"] { background:transparent !important; padding:1rem 0 !important; }
[data-testid="stDataFrame"] { border-radius:10px; overflow:hidden; }
hr { border-color:var(--border) !important; margin:1.2rem 0 !important; }
.sidebar-section { font-family:'Barlow Condensed',sans-serif !important; font-size:0.7rem; font-weight:700;
    letter-spacing:0.15em; text-transform:uppercase; color:var(--accent);
    margin:1rem 0 0.5rem; padding-bottom:0.3rem; border-bottom:1px solid var(--border); }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────
def safe_pct(num, den):
    """Divisione protetta che restituisce 0.0 invece di inf/NaN."""
    try:
        if den is None or pd.isna(den) or den == 0:
            return 0.0
        return round(float(num) / float(den) * 100.0, 1)
    except Exception:
        return 0.0


def safe_str(v, default="—"):
    if v is None:
        return default
    try:
        if pd.isna(v):
            return default
    except Exception:
        pass
    return str(v)


def latin1(s):
    """
    Sanitizza testo per PDF FPDF (font core Helvetica = latin-1).
    Sostituisce caratteri non supportati con equivalenti ASCII.
    """
    if s is None:
        return ""
    s = str(s)
    repl = {
        "–": "-", "—": "-", "−": "-",
        "“": '"', "”": '"', "„": '"', "‟": '"',
        "‘": "'", "’": "'", "‚": ",", "‛": "'",
        "…": "...", "•": "*", "·": "-",
        "→": "->", "←": "<-", "↑": "^", "↓": "v",
        "▲": "^", "▼": "v", "♾️": "inf", "✓": "v", "✗": "x",
        "€": "EUR", "£": "GBP",
        "Δ": "Delta", "%": "%",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s.encode("latin-1", errors="replace").decode("latin-1")


# ─────────────────────────────────────────────────────────────────
#  FOREVER-ADJUSTED HELPERS (Proposta 4: Standard / Adjusted / Confronto)
# ─────────────────────────────────────────────────────────────────
# Modalità calcolo:
#   "standard"  : Gross/TAM, Net/TAM (calcolo classico Wind3)
#   "adjusted"  : (Gross+Forever)/(TAM+Forever), (Net+Forever)/(TAM+Forever)
#                 → realtà del lavoro: il telefono Forever conta come venduto
#                   con reload effettiva sopra la baseline TAM
#   "confronto" : entrambi affiancati nei tile e nelle cards
#
# Le Forever sono "premio" sopra la baseline TAM perché sono telefoni che hanno
# superato il mese di osservazione e contano come vendita consolidata.
# Esempio: TAM=100, Gross=50, Forever=10 → adjusted: 60/110 = 54.5%

def gross_pct_calc(gross, tam, forever, mode="standard"):
    """Calcolo Gross% in base alla modalità."""
    if mode == "adjusted":
        return safe_pct((gross or 0) + (forever or 0), (tam or 0) + (forever or 0))
    return safe_pct(gross, tam)


def net_pct_calc(net, tam, forever, mode="standard"):
    """Calcolo Net% in base alla modalità."""
    if mode == "adjusted":
        return safe_pct((net or 0) + (forever or 0), (tam or 0) + (forever or 0))
    return safe_pct(net, tam)


def fmt_compare(std_val, adj_val, mode):
    """Formatta un valore in base alla modalità (per UI)."""
    if mode == "confronto":
        return f"{std_val}% / {adj_val}%"
    if mode == "adjusted":
        return f"{adj_val}%"
    return f"{std_val}%"


# ─────────────────────────────────────────────────────────────────
#  CORE LOGIC
# ─────────────────────────────────────────────────────────────────
REQUIRED_COLS = [
    "SHOP_CODE", "STORE", "COMPANY_NAME", "CITY", "PROVINCE_CODE",
    "STORE_TYPE", "REGION", "AMBASSADOR", "AREA_MANAGER", "DISTRICT_MANAGER",
    "TAM", "SR_PLUS_GROSS_SALES", "SR_PLUS_NET_SALES",
    "Activeforever", "TotalStore", "MONTH", "YEAR",
]

OPTIONAL_COLS = ["STORE_ADDRESS"]

# Colonne aggiuntive (presenti nei file Wind3 con 76 colonne)
FASCE_PREZZO_COLS = [
    "TAM_0_150", "SR_PLUS_GROSS_SALES_0_150",
    "TAM_150_300", "SR_PLUS_GROSS_SALES_150_300",
    "TAM_300_700", "SR_PLUS_GROSS_SALES_300_700",
    "TAM_700_1200", "SR_PLUS_GROSS_SALES_700_1200",
    "TAM_1200_1600", "SR_PLUS_GROSS_SALES_1200_1600",
    "TAM_>1600", "SR_PLUS_GROSS_SALES_>1600",
]

RELOAD_PLUS_COLS = [
    "PLUS_GROSS_SALES", "PLUS_NET_SALES", "TAM_ELIGIBLE_PLUS",
    "SR_GROSS_SALES", "SR_NET_SALES",
]

RELOAD_FOREVER_COLS = [
    "R4_BASIC_GROSS_SALES", "R4_BASIC_NET_SALES",
    "R4_PREMIUM_GROSS_SALES", "R4_PREMIUM_NET_SALES",
]

RELOAD_EU_COLS = [
    "RELOADEU_PLUS_GROSS_SALES", "RELOADEU_PLUS_NET_SALES",
    "RELOADEU_SMART_GROSS_SALES", "RELOADEU_SMART_NET_SALES",
    "RELOADEU_OPEN_GROSS_SALES", "RELOADEU_OPEN_NET_SALES",
]

EXCHANGE_COLS = ["EXCHANGE_SALES", "EXCHANGE_SALES_VALUABLE", "EXCHANGE_SALES_NO_VALUABLE"]

ALL_OPTIONAL_NUMERIC = (
    FASCE_PREZZO_COLS + RELOAD_PLUS_COLS + RELOAD_FOREVER_COLS
    + RELOAD_EU_COLS + EXCHANGE_COLS
)

# ─────────────────────────────────────────────────────────────────
#  FASCE TELEFONO E GETTONI RELOAD (PPTX Aprile 2026, slide 16/32)
# ─────────────────────────────────────────────────────────────────
# Le 5 fasce ufficiali del PPTX (1° fascia <150€ ... 5° fascia >1200€)
# Mappiamo le 6 fasce del file Excel verso le 5 del PPTX
#
# Excel:  0-150 / 150-300 / 300-700 / 700-1200 / 1200-1600 / >1600
# PPTX:   <150  / 150-300 / 300-700 / 700-1200 / >1200
#
# → 1200-1600 e >1600 vengono entrambi mappati alla 5° fascia PPTX
FASCE_PPTX = [
    {"key": "0_150",     "label": "1° (<150€)",      "tam_col": "TAM_0_150",      "gross_col": "SR_PLUS_GROSS_SALES_0_150"},
    {"key": "150_300",   "label": "2° (150-300€)",   "tam_col": "TAM_150_300",    "gross_col": "SR_PLUS_GROSS_SALES_150_300"},
    {"key": "300_700",   "label": "3° (300-700€)",   "tam_col": "TAM_300_700",    "gross_col": "SR_PLUS_GROSS_SALES_300_700"},
    {"key": "700_1200",  "label": "4° (700-1200€)",  "tam_col": "TAM_700_1200",   "gross_col": "SR_PLUS_GROSS_SALES_700_1200"},
    {"key": "1200_1600", "label": "5a (1200-1600€)", "tam_col": "TAM_1200_1600",  "gross_col": "SR_PLUS_GROSS_SALES_1200_1600"},
    {"key": ">1600",     "label": "5b (>1600€)",     "tam_col": "TAM_>1600",      "gross_col": "SR_PLUS_GROSS_SALES_>1600"},
]

# Gettoni Smartphone Reload (€ per unità) — PPTX Aprile 2026, slide 16/32
# Per ogni fascia, gettoni per: Reload, Reload EU, Forever Basic, Forever Premium, Reload Plus, Reload Plus EU
# Le 5e fasce (1200-1600 e >1600) usano lo stesso gettone della 5° fascia PPTX
GETTONI_RELOAD_DEFAULT = {
    "0_150":     {"reload": 5.0,  "reload_eu": 6.0,  "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 0.0,  "reload_plus_eu": 0.0},
    "150_300":   {"reload": 6.0,  "reload_eu": 7.0,  "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 0.0,  "reload_plus_eu": 0.0},
    "300_700":   {"reload": 7.0,  "reload_eu": 9.0,  "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 7.0,  "reload_plus_eu": 10.0},
    "700_1200":  {"reload": 9.0,  "reload_eu": 11.0, "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 9.0,  "reload_plus_eu": 12.0},
    "1200_1600": {"reload": 11.0, "reload_eu": 14.0, "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 11.0, "reload_plus_eu": 15.0},
    ">1600":     {"reload": 11.0, "reload_eu": 14.0, "forever_basic": 5.0, "forever_premium": 7.0, "reload_plus": 11.0, "reload_plus_eu": 15.0},
}

# Soglie Net Attachment Rate per moltiplicatore X2/X3/X4 (PPTX slide 16/32)
# In base al canale commerciale del PDV
SOGLIE_AR_DEFAULT = {
    "Franchising":  {"x2": 40, "x3": 60, "x4": 75},
    "W3R":          {"x2": 55, "x3": 65, "x4": 75},
    "Top Quality":  {"x2": 25, "x3": 40, "x4": 55},
    "Dealer":       {"x2": 25, "x3": 40, "x4": 55},
    "W3SP":         {"x2": 15, "x3": 30, "x4": 50},
    "Corner":       {"x2": 35, "x3": 45, "x4": 55},  # da slide 59
}

# Mapping default STORE_TYPE Excel → canale PPTX (l'utente può sovrascriverlo)
STORE_TYPE_TO_CANALE_DEFAULT = {
    "Franchising":  "Franchising",
    "Owned Stores": "W3R",
    "Dealer":       "Dealer",
    "Large Chain":  "Top Quality",
    "Other":        "W3SP",
}


def load_and_filter(file_bytes, cfg):
    """
    Carica il file Excel e filtra per ambassador/regioni.
    Bug fix v4:
      - filtro AND (non OR) tra ambassador e regioni
      - 'latest' calcolato DOPO il filtro per zona
      - controllo colonne mancanti
      - colonne opzionali create vuote se assenti
    """
    df = pd.read_excel(file_bytes, sheet_name=cfg["sheet_name"])

    # Check colonne obbligatorie
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel file: {', '.join(missing)}")

    # Colonne opzionali stringa
    for c in OPTIONAL_COLS:
        if c not in df.columns:
            df[c] = ""

    # Colonne opzionali numeriche (fasce, plus, R4, EU, exchange) → 0 se assenti
    for c in ALL_OPTIONAL_NUMERIC:
        if c not in df.columns:
            df[c] = 0
        else:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    valid = df[df["SHOP_CODE"].notna()].copy()

    # Applica filtro zona PRIMA di calcolare 'latest'
    amb_filter = cfg.get("ambassador_filter") or []
    reg_filter = cfg.get("regioni_filter") or []
    filter_mode = cfg.get("filter_mode", "AND")  # "AND" | "OR" | "AMB" | "REG"

    if filter_mode == "AND":
        m = pd.Series(True, index=valid.index)
        if amb_filter:
            m &= valid["AMBASSADOR"].isin(amb_filter)
        if reg_filter:
            m &= valid["REGION"].isin(reg_filter)
    elif filter_mode == "OR":
        m = pd.Series(False, index=valid.index)
        if amb_filter:
            m |= valid["AMBASSADOR"].isin(amb_filter)
        if reg_filter:
            m |= valid["REGION"].isin(reg_filter)
        if not amb_filter and not reg_filter:
            m = pd.Series(True, index=valid.index)
    elif filter_mode == "AMB":
        m = valid["AMBASSADOR"].isin(amb_filter) if amb_filter else pd.Series(True, index=valid.index)
    else:  # REG
        m = valid["REGION"].isin(reg_filter) if reg_filter else pd.Series(True, index=valid.index)

    zone = valid[m].copy()

    if zone.empty:
        raise ValueError(
            "Nessun dato per la zona selezionata. "
            "Controlla Ambassador / Regioni / modalità filtro."
        )

    # 'latest' calcolato sulla zona filtrata
    latest = zone.sort_values(["YEAR", "MONTH"], ascending=False).iloc[0]
    month, year = int(latest["MONTH"]), int(latest["YEAR"])

    current = zone[(zone["MONTH"] == month) & (zone["YEAR"] == year)].copy()

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev = zone[(zone["MONTH"] == prev_month) & (zone["YEAR"] == prev_year)].copy()

    meta = {
        "month": month, "year": year,
        "prev_month": prev_month, "prev_year": prev_year,
    }
    return current, prev, meta


def compute_kpis(df, df_prev, cfg):
    """Aggrega per DM e calcola KPI + status. Supporta modalità Standard/Adjusted/Confronto."""
    if df.empty:
        return pd.DataFrame()

    calc_mode = cfg.get("calc_mode", "standard")

    agg = df.groupby(["REGION", "AREA_MANAGER", "DISTRICT_MANAGER"], dropna=False).agg(
        TAM=("TAM", "sum"),
        Gross_Sales=("SR_PLUS_GROSS_SALES", "sum"),
        Net_Sales=("SR_PLUS_NET_SALES", "sum"),
        Forever_Active=("Activeforever", "sum"),
        Total_Stores=("TotalStore", "sum"),
        N_Negozi=("SHOP_CODE", "count"),
    ).reset_index()

    # Calcoli STANDARD (sempre presenti)
    agg["Gross_%_std"] = agg.apply(lambda r: safe_pct(r["Gross_Sales"], r["TAM"]), axis=1)
    agg["Net_%_std"] = agg.apply(lambda r: safe_pct(r["Net_Sales"], r["TAM"]), axis=1)
    # Calcoli ADJUSTED (Forever sopra baseline)
    agg["Gross_%_adj"] = agg.apply(
        lambda r: gross_pct_calc(r["Gross_Sales"], r["TAM"], r["Forever_Active"], "adjusted"), axis=1)
    agg["Net_%_adj"] = agg.apply(
        lambda r: net_pct_calc(r["Net_Sales"], r["TAM"], r["Forever_Active"], "adjusted"), axis=1)

    # Le colonne "ufficiali" che il resto dell'app userà cambiano in base alla modalità
    if calc_mode == "adjusted":
        agg["Gross_%"] = agg["Gross_%_adj"]
        agg["Net_%"] = agg["Net_%_adj"]
    else:
        # standard e confronto usano lo standard come riferimento per status/sort
        agg["Gross_%"] = agg["Gross_%_std"]
        agg["Net_%"] = agg["Net_%_std"]

    agg["Forever_%"] = agg.apply(lambda r: safe_pct(r["Forever_Active"], r["Total_Stores"]), axis=1)

    if df_prev is not None and not df_prev.empty:
        pa = df_prev.groupby("DISTRICT_MANAGER", dropna=False).agg(
            TAM_p=("TAM", "sum"),
            Gross_p=("SR_PLUS_GROSS_SALES", "sum"),
            Net_p=("SR_PLUS_NET_SALES", "sum"),
            Forever_p=("Activeforever", "sum"),
        ).reset_index()
        if calc_mode == "adjusted":
            pa["Gross_%_p"] = pa.apply(
                lambda r: gross_pct_calc(r["Gross_p"], r["TAM_p"], r["Forever_p"], "adjusted"), axis=1)
            pa["Net_%_p"] = pa.apply(
                lambda r: net_pct_calc(r["Net_p"], r["TAM_p"], r["Forever_p"], "adjusted"), axis=1)
        else:
            pa["Gross_%_p"] = pa.apply(lambda r: safe_pct(r["Gross_p"], r["TAM_p"]), axis=1)
            pa["Net_%_p"] = pa.apply(lambda r: safe_pct(r["Net_p"], r["TAM_p"]), axis=1)
        agg = agg.merge(pa[["DISTRICT_MANAGER", "Gross_%_p", "Net_%_p"]], on="DISTRICT_MANAGER", how="left")
        agg["Gross_Δ"] = (agg["Gross_%"] - agg["Gross_%_p"]).round(1)
        agg["Net_Δ"] = (agg["Net_%"] - agg["Net_%_p"]).round(1)
    else:
        agg["Gross_Δ"] = pd.NA
        agg["Net_Δ"] = pd.NA

    sg = cfg["soglia_critica_gross"]
    sn = cfg["soglia_critica_net"]
    sa_g = cfg.get("soglia_attenzione_gross", sg + 20)
    sa_n = cfg.get("soglia_attenzione_net", sn + 20)
    forever_zero_critico = cfg.get("forever_zero_critico", True)

    def status(row):
        alerts = []
        is_critical = False
        if row["Gross_%"] < sg:
            alerts.append(f"Gross {row['Gross_%']}% < {sg}%")
            is_critical = True
        if row["Net_%"] < sn:
            alerts.append(f"Net {row['Net_%']}% < {sn}%")
            is_critical = True
        if row["Forever_Active"] == 0:
            if forever_zero_critico:
                alerts.append("Forever = 0")
                is_critical = True
            else:
                alerts.append("Forever = 0 (warn)")
        if not alerts:
            return "ok", []
        if is_critical:
            return "critico", alerts
        # attenzione: sopra critica ma sotto attenzione
        if row["Gross_%"] < sa_g or row["Net_%"] < sa_n:
            return "attenzione", alerts
        return "attenzione", alerts

    statuses = agg.apply(status, axis=1)
    agg["status"] = [s[0] for s in statuses]
    agg["alerts"] = [s[1] for s in statuses]

    def tl(d):
        if d is None or pd.isna(d):
            return "—"
        if d > 0:
            return f"▲ +{d}%"
        if d < 0:
            return f"▼ {d}%"
        return "= 0%"

    agg["Gross_Trend"] = agg["Gross_Δ"].apply(tl)
    agg["Net_Trend"] = agg["Net_Δ"].apply(tl)
    return agg.sort_values("Gross_%", ascending=False).reset_index(drop=True)


def build_hierarchy(df):
    """
    Costruisce mappa gerarchica ORDINATA: AM → {DM → [stores]}.
    - AM ordinati per Gross% decrescente
    - DM dentro ogni AM ordinati per Gross% decrescente
    - Store dentro ogni DM ordinati per Gross% decrescente
    Restituisce list di tuple per rispettare l'ordine (i dict Python 3.7+ mantengono l'ordine
    di inserimento, ma usiamo list-of-tuples per chiarezza).
    """
    tree = {}
    for am, am_df in df.groupby("AREA_MANAGER", dropna=False):
        am_key = safe_str(am, "(senza AM)")
        am_tam = int(am_df["TAM"].sum())
        am_gross = int(am_df["SR_PLUS_GROSS_SALES"].sum())
        am_net = int(am_df["SR_PLUS_NET_SALES"].sum())
        am_data = {
            "regions": sorted(am_df["REGION"].dropna().unique().tolist()),
            "n_dm": am_df["DISTRICT_MANAGER"].nunique(),
            "n_stores": len(am_df),
            "tam": am_tam,
            "gross": am_gross,
            "net": am_net,
            "gross_pct": safe_pct(am_gross, am_tam),
            "net_pct": safe_pct(am_net, am_tam),
            "dms": {},
        }
        # DM list con metriche per ordinamento
        dm_list = []
        for dm, dm_df in am_df.groupby("DISTRICT_MANAGER", dropna=False):
            dm_key = safe_str(dm, "(senza DM)")
            dm_tam = int(dm_df["TAM"].sum())
            dm_gross = int(dm_df["SR_PLUS_GROSS_SALES"].sum())
            dm_net = int(dm_df["SR_PLUS_NET_SALES"].sum())
            stores = []
            for _, s in dm_df.iterrows():
                s_tam = int(s.get("TAM", 0) or 0)
                s_gross = int(s.get("SR_PLUS_GROSS_SALES", 0) or 0)
                s_net = int(s.get("SR_PLUS_NET_SALES", 0) or 0)
                stores.append({
                    "store": safe_str(s.get("STORE")),
                    "company": safe_str(s.get("COMPANY_NAME")),
                    "city": safe_str(s.get("CITY")),
                    "province": safe_str(s.get("PROVINCE_CODE")),
                    "type": safe_str(s.get("STORE_TYPE")),
                    "tam": s_tam,
                    "gross": s_gross,
                    "net": s_net,
                    "gross_pct": safe_pct(s_gross, s_tam),
                    "net_pct": safe_pct(s_net, s_tam),
                    "forever": int(s.get("Activeforever", 0) or 0),
                })
            # Ordina store per Gross% decrescente
            stores.sort(key=lambda x: x["gross_pct"], reverse=True)
            dm_list.append((dm_key, {
                "n_stores": len(stores),
                "tam": dm_tam,
                "gross": dm_gross,
                "net": dm_net,
                "gross_pct": safe_pct(dm_gross, dm_tam),
                "net_pct": safe_pct(dm_net, dm_tam),
                "regions": sorted(dm_df["REGION"].dropna().unique().tolist()),
                "stores": stores,
            }))
        # Ordina DM per Gross% decrescente
        dm_list.sort(key=lambda x: x[1]["gross_pct"], reverse=True)
        am_data["dms"] = dict(dm_list)
        tree[am_key] = am_data

    # Ordina AM per Gross% decrescente
    tree_sorted = dict(sorted(tree.items(), key=lambda x: x[1]["gross_pct"], reverse=True))
    return tree_sorted


def analyze_gross_net(df, soglia_gross, soglia_net, sa_gross, sa_net):
    """
    Analisi statistica completa Gross e Net sull'intera zona.
    Restituisce un dict con tutte le metriche aggregate.
    """
    if df.empty:
        return None

    tam = float(df["TAM"].sum())
    gross = float(df["SR_PLUS_GROSS_SALES"].sum())
    net = float(df["SR_PLUS_NET_SALES"].sum())

    gross_pct = safe_pct(gross, tam)
    net_pct = safe_pct(net, tam)
    net_on_gross = safe_pct(net, gross)  # tasso di conversione gross→net

    # Per-store metrics
    store_df = df.copy()
    store_df["g_pct"] = store_df.apply(lambda r: safe_pct(r["SR_PLUS_GROSS_SALES"], r["TAM"]), axis=1)
    store_df["n_pct"] = store_df.apply(lambda r: safe_pct(r["SR_PLUS_NET_SALES"], r["TAM"]), axis=1)

    # Distribuzione per fasce Gross
    fasce_gross = {
        "🔴 Critici (< {}%)".format(soglia_gross): int((store_df["g_pct"] < soglia_gross).sum()),
        "🟠 Attenzione ({}-{}%)".format(soglia_gross, sa_gross): int(
            ((store_df["g_pct"] >= soglia_gross) & (store_df["g_pct"] < sa_gross)).sum()),
        "🟢 OK (≥ {}%)".format(sa_gross): int((store_df["g_pct"] >= sa_gross).sum()),
    }
    fasce_net = {
        "🔴 Critici (< {}%)".format(soglia_net): int((store_df["n_pct"] < soglia_net).sum()),
        "🟠 Attenzione ({}-{}%)".format(soglia_net, sa_net): int(
            ((store_df["n_pct"] >= soglia_net) & (store_df["n_pct"] < sa_net)).sum()),
        "🟢 OK (≥ {}%)".format(sa_net): int((store_df["n_pct"] >= sa_net).sum()),
    }

    # Top / Bottom per Gross e Net (top 5)
    top_gross = store_df.nlargest(5, "g_pct")[["STORE", "DISTRICT_MANAGER", "g_pct", "n_pct", "TAM"]].to_dict("records")
    bot_gross = store_df[store_df["TAM"] > 0].nsmallest(5, "g_pct")[["STORE", "DISTRICT_MANAGER", "g_pct", "n_pct", "TAM"]].to_dict("records")
    top_net = store_df.nlargest(5, "n_pct")[["STORE", "DISTRICT_MANAGER", "g_pct", "n_pct", "TAM"]].to_dict("records")
    bot_net = store_df[store_df["TAM"] > 0].nsmallest(5, "n_pct")[["STORE", "DISTRICT_MANAGER", "g_pct", "n_pct", "TAM"]].to_dict("records")

    # Gap rispetto alla soglia di attenzione (quanto manca per portare tutti in zona "ok")
    needed_gross = max(0.0, sa_gross / 100.0 * tam - gross)
    needed_net = max(0.0, sa_net / 100.0 * tam - net)

    # Per tipo store
    by_type = df.groupby("STORE_TYPE", dropna=False).agg(
        n=("SHOP_CODE", "count"),
        tam=("TAM", "sum"),
        gross=("SR_PLUS_GROSS_SALES", "sum"),
        net=("SR_PLUS_NET_SALES", "sum"),
    ).reset_index()
    by_type["gross_pct"] = by_type.apply(lambda r: safe_pct(r["gross"], r["tam"]), axis=1)
    by_type["net_pct"] = by_type.apply(lambda r: safe_pct(r["net"], r["tam"]), axis=1)
    by_type["share_gross"] = by_type.apply(lambda r: safe_pct(r["gross"], gross), axis=1)
    by_type = by_type.sort_values("gross", ascending=False)

    return {
        "tam": int(tam), "gross": int(gross), "net": int(net),
        "gross_pct": gross_pct, "net_pct": net_pct,
        "net_on_gross": net_on_gross,
        "n_stores": len(store_df),
        "n_stores_zero_tam": int((store_df["TAM"] == 0).sum()),
        "fasce_gross": fasce_gross, "fasce_net": fasce_net,
        "top_gross": top_gross, "bot_gross": bot_gross,
        "top_net": top_net, "bot_net": bot_net,
        "needed_gross": int(needed_gross),
        "needed_net": int(needed_net),
        "by_type": by_type,
        "avg_gross_per_store": gross_pct,
        "median_gross_pct": float(store_df["g_pct"].median()),
        "median_net_pct": float(store_df["n_pct"].median()),
    }


def generate_message(row, meta):
    mese_str = datetime(meta["year"], meta["month"], 1).strftime("%B %Y").capitalize()

    def ts(d):
        if d is None or (isinstance(d, float) and pd.isna(d)):
            return ""
        if d > 0:
            return f"▲ +{d}%"
        if d < 0:
            return f"▼ {d}%"
        return "= 0%"

    fi = "🔴" if row["Forever_Active"] == 0 else "♾️"
    alerts = row.get("alerts", [])
    ab = "\n".join([f"⚠️ {a}" for a in alerts]) if alerts else "✅ Tutti i KPI nella norma"
    return f"""📊 *AVANZAMENTO RELOAD – {mese_str}*
👤 *{row['DISTRICT_MANAGER']}*
📍 {row['REGION']} | AM: {row['AREA_MANAGER']}
─────────────────────────
🔵 *Gross (Lordo):*   {int(row['Gross_Sales'])} / {int(row['TAM'])}  →  *{row['Gross_%']}%*  {ts(row.get('Gross_Δ'))}
🟢 *Net (Netto):*     {int(row['Net_Sales'])} / {int(row['TAM'])}  →  *{row['Net_%']}%*  {ts(row.get('Net_Δ'))}
{fi}  *Reload Forever:*  {int(row['Forever_Active'])} / {int(row['Total_Stores'])} store  →  *{row['Forever_%']}%*
─────────────────────────
🏪 Negozi: {int(row['N_Negozi'])}
{ab}"""


# ─────────────────────────────────────────────────────────────────
#  FASE 1: AGGREGAZIONE PER RAGIONE SOCIALE
# ─────────────────────────────────────────────────────────────────
def compute_rs_aggregates(df, calc_mode="standard"):
    """
    Aggrega tutti i dati per COMPANY_NAME (Ragione Sociale).
    Restituisce un DataFrame con una riga per RS, con tutti i totali utili
    per analisi e calcolo compensi.
    """
    if df.empty:
        return pd.DataFrame()

    # Aggregati base
    agg_dict = {
        "N_PDV":           ("SHOP_CODE", "count"),
        "TAM":             ("TAM", "sum"),
        "Gross":           ("SR_PLUS_GROSS_SALES", "sum"),
        "Net":             ("SR_PLUS_NET_SALES", "sum"),
        "Forever_Active":  ("Activeforever", "sum"),
        "Total_Stores":    ("TotalStore", "sum"),
    }
    # Aggregati opzionali se le colonne ci sono
    extra_cols = {
        "TAM_0_150": "tam_0_150", "TAM_150_300": "tam_150_300",
        "TAM_300_700": "tam_300_700", "TAM_700_1200": "tam_700_1200",
        "TAM_1200_1600": "tam_1200_1600", "TAM_>1600": "tam_>1600",
        "SR_PLUS_GROSS_SALES_0_150": "g_0_150", "SR_PLUS_GROSS_SALES_150_300": "g_150_300",
        "SR_PLUS_GROSS_SALES_300_700": "g_300_700", "SR_PLUS_GROSS_SALES_700_1200": "g_700_1200",
        "SR_PLUS_GROSS_SALES_1200_1600": "g_1200_1600", "SR_PLUS_GROSS_SALES_>1600": "g_>1600",
        "PLUS_GROSS_SALES": "plus_gross", "PLUS_NET_SALES": "plus_net",
        "R4_BASIC_GROSS_SALES": "r4_basic_gross", "R4_BASIC_NET_SALES": "r4_basic_net",
        "R4_PREMIUM_GROSS_SALES": "r4_premium_gross", "R4_PREMIUM_NET_SALES": "r4_premium_net",
        "RELOADEU_PLUS_GROSS_SALES": "reload_eu_plus_gross", "RELOADEU_PLUS_NET_SALES": "reload_eu_plus_net",
        "RELOADEU_SMART_GROSS_SALES": "reload_eu_smart_gross", "RELOADEU_SMART_NET_SALES": "reload_eu_smart_net",
        "EXCHANGE_SALES": "exchange",
    }
    for src, dst in extra_cols.items():
        if src in df.columns:
            agg_dict[dst] = (src, "sum")

    rs = df.groupby("COMPANY_NAME", dropna=False).agg(**agg_dict).reset_index()

    # Anagrafica derivata: AM, DM, regioni, store types
    anagrafica = df.groupby("COMPANY_NAME", dropna=False).agg(
        AM=("AREA_MANAGER", lambda s: ", ".join(sorted(set(safe_str(x) for x in s if pd.notna(x))))),
        DM=("DISTRICT_MANAGER", lambda s: ", ".join(sorted(set(safe_str(x) for x in s if pd.notna(x))))),
        Regioni=("REGION", lambda s: ", ".join(sorted(set(safe_str(x) for x in s if pd.notna(x))))),
        Tipi_Store=("STORE_TYPE", lambda s: ", ".join(sorted(set(safe_str(x) for x in s if pd.notna(x))))),
        Province=("PROVINCE_CODE", lambda s: ", ".join(sorted(set(safe_str(x) for x in s if pd.notna(x))))),
    ).reset_index()
    rs = rs.merge(anagrafica, on="COMPANY_NAME", how="left")

    # KPI calcolati
    rs["Gross_%_std"] = rs.apply(lambda r: safe_pct(r["Gross"], r["TAM"]), axis=1)
    rs["Net_%_std"]   = rs.apply(lambda r: safe_pct(r["Net"], r["TAM"]), axis=1)
    rs["Gross_%_adj"] = rs.apply(
        lambda r: gross_pct_calc(r["Gross"], r["TAM"], r["Forever_Active"], "adjusted"), axis=1)
    rs["Net_%_adj"]   = rs.apply(
        lambda r: net_pct_calc(r["Net"], r["TAM"], r["Forever_Active"], "adjusted"), axis=1)

    if calc_mode == "adjusted":
        rs["Gross_%"] = rs["Gross_%_adj"]
        rs["Net_%"] = rs["Net_%_adj"]
    else:
        rs["Gross_%"] = rs["Gross_%_std"]
        rs["Net_%"] = rs["Net_%_std"]

    rs["Forever_%"] = rs.apply(lambda r: safe_pct(r["Forever_Active"], r["Total_Stores"]), axis=1)

    return rs.sort_values("Gross", ascending=False).reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────
#  FASE 2: CALCOLATORE COMPENSI RELOAD
# ─────────────────────────────────────────────────────────────────
def determina_canale(store_type, mapping=None):
    """Mappa STORE_TYPE Excel → canale PPTX usando mapping (default + override)."""
    m = dict(STORE_TYPE_TO_CANALE_DEFAULT)
    if mapping:
        m.update(mapping)
    return m.get(safe_str(store_type), "Dealer")


def calcola_moltiplicatore_ar(net_ar_pct, canale, soglie_ar=None):
    """
    Calcola il moltiplicatore Smartphone Reload (X1/X2/X3/X4)
    in base al Net Attachment Rate e al canale.
    """
    soglie = soglie_ar or SOGLIE_AR_DEFAULT
    s = soglie.get(canale, soglie.get("Dealer", {"x2": 25, "x3": 40, "x4": 55}))
    if net_ar_pct >= s["x4"]:
        return 4
    if net_ar_pct >= s["x3"]:
        return 3
    if net_ar_pct >= s["x2"]:
        return 2
    return 1


def compute_compensi_reload(df, gettoni=None, soglie_ar=None,
                            store_type_mapping=None, calc_mode="standard"):
    """
    Calcola il compenso teorico Smartphone Reload per ogni Ragione Sociale.

    Versione Fase 4: Plus distribuito per fascia in base alla quota del Gross della fascia.
    Reload EU Plus segue le fasce 3-4-5 (>=300€), Reload EU Smart segue tutte le fasce.

    Per ogni RS:
    1. Aggrega vendite per fascia di prezzo telefono (6 fasce Excel → 5 PPTX)
    2. Distingue tipologia reload: base, Forever Basic/Premium, Plus, EU
    3. Distribuisce Plus e EU sulle fasce in base al peso del Gross di fascia
    4. Determina il canale dominante della RS dai suoi store
    5. Calcola Net Attachment Rate per RS
    6. Applica gettoni × vendite, somma per fascia/tipo
    7. Moltiplica per X1/X2/X3/X4 in base ad AR e canale
    """
    if df.empty:
        return pd.DataFrame()

    gettoni = gettoni or GETTONI_RELOAD_DEFAULT
    soglie_ar = soglie_ar or SOGLIE_AR_DEFAULT

    rows = []
    for company, sub in df.groupby("COMPANY_NAME", dropna=False):
        # Anagrafica RS
        n_pdv = len(sub)
        canali_rs = [determina_canale(t, store_type_mapping) for t in sub["STORE_TYPE"]]
        canale_dom = max(set(canali_rs), key=canali_rs.count) if canali_rs else "Dealer"

        # === Vendite per fascia ===
        gross_per_fascia = {}
        tam_per_fascia = {}
        for fascia in FASCE_PPTX:
            tam_per_fascia[fascia["key"]] = float(sub.get(fascia["tam_col"], pd.Series([0])).sum() or 0)
            gross_per_fascia[fascia["key"]] = float(sub.get(fascia["gross_col"], pd.Series([0])).sum() or 0)
        gross_tot_fasce = sum(gross_per_fascia.values())

        # === Reload base (per fascia) ===
        compenso_base = 0.0
        dettaglio_fasce = {}
        for fascia in FASCE_PPTX:
            g_fascia = gross_per_fascia[fascia["key"]]
            tam_fascia = tam_per_fascia[fascia["key"]]
            gettone_base = gettoni.get(fascia["key"], {}).get("reload", 0.0)
            comp_f = g_fascia * gettone_base
            compenso_base += comp_f
            dettaglio_fasce[fascia["key"]] = {
                "tam": int(tam_fascia), "gross": int(g_fascia),
                "gettone": gettone_base, "compenso": comp_f,
            }

        # === Reload Plus (Fase 4: distribuito per fascia ≥ 300€) ===
        plus_gross = float(sub.get("PLUS_GROSS_SALES", pd.Series([0])).sum() or 0)
        compenso_plus = 0.0
        # Plus esiste solo per fasce ≥ 300€ → distribuisco proporzionalmente al Gross di queste fasce
        plus_eligible = ["300_700", "700_1200", "1200_1600", ">1600"]
        gross_eligible_plus = sum(gross_per_fascia[k] for k in plus_eligible)
        if gross_eligible_plus > 0 and plus_gross > 0:
            for k in plus_eligible:
                quota = gross_per_fascia[k] / gross_eligible_plus
                vendite_plus_fascia = plus_gross * quota
                gettone_plus = gettoni.get(k, {}).get("reload_plus", 0.0)
                compenso_plus += vendite_plus_fascia * gettone_plus
        else:
            # fallback: gettone medio se non riesco a distribuire
            compenso_plus = plus_gross * 8.5

        # === Reload EU (Plus + Smart) ===
        reload_eu_plus = float(sub.get("RELOADEU_PLUS_GROSS_SALES", pd.Series([0])).sum() or 0)
        reload_eu_smart = float(sub.get("RELOADEU_SMART_GROSS_SALES", pd.Series([0])).sum() or 0)
        compenso_eu = 0.0
        # EU Plus → distribuito sulle fasce >=300 col gettone reload_plus_eu
        if gross_eligible_plus > 0 and reload_eu_plus > 0:
            for k in plus_eligible:
                quota = gross_per_fascia[k] / gross_eligible_plus
                vendite_eu_plus_fascia = reload_eu_plus * quota
                gettone = gettoni.get(k, {}).get("reload_plus_eu", 0.0)
                compenso_eu += vendite_eu_plus_fascia * gettone
        else:
            compenso_eu += reload_eu_plus * 11.0
        # EU Smart → distribuito su tutte le fasce col gettone reload_eu
        if gross_tot_fasce > 0 and reload_eu_smart > 0:
            for fascia in FASCE_PPTX:
                quota = gross_per_fascia[fascia["key"]] / gross_tot_fasce
                vendite_eu_smart_fascia = reload_eu_smart * quota
                gettone = gettoni.get(fascia["key"], {}).get("reload_eu", 0.0)
                compenso_eu += vendite_eu_smart_fascia * gettone
        else:
            compenso_eu += reload_eu_smart * 8.0

        # === Forever Basic e Premium (gettone fisso 5€/7€ indipendente dalla fascia) ===
        forever_basic_gross = float(sub.get("R4_BASIC_GROSS_SALES", pd.Series([0])).sum() or 0)
        forever_premium_gross = float(sub.get("R4_PREMIUM_GROSS_SALES", pd.Series([0])).sum() or 0)
        compenso_forever_basic = forever_basic_gross * 5.0
        compenso_forever_premium = forever_premium_gross * 7.0

        # === Reload Exchange (gettone fisso 5€) ===
        exchange = float(sub.get("EXCHANGE_SALES", pd.Series([0])).sum() or 0)
        compenso_exchange = exchange * 5.0

        # === KPI per RS ===
        tam_tot = float(sub["TAM"].sum())
        gross_tot = float(sub["SR_PLUS_GROSS_SALES"].sum())
        net_tot = float(sub["SR_PLUS_NET_SALES"].sum())
        forever_active = float(sub["Activeforever"].sum())

        # Net Attachment Rate (formula PPTX slide 16):
        # (Reload Net + Plus Net + Forever attive) / (TAM + Forever attive)
        plus_net = float(sub.get("PLUS_NET_SALES", pd.Series([0])).sum() or 0)
        net_ar = safe_pct(net_tot + plus_net + forever_active, tam_tot + forever_active)

        moltiplicatore = calcola_moltiplicatore_ar(net_ar, canale_dom, soglie_ar)

        compenso_subtotal = (compenso_base + compenso_forever_basic + compenso_forever_premium
                             + compenso_plus + compenso_eu + compenso_exchange)
        compenso_totale = compenso_subtotal * moltiplicatore

        rows.append({
            "Ragione Sociale": safe_str(company),
            "Canale": canale_dom,
            "N PDV": n_pdv,
            "TAM": int(tam_tot),
            "Gross": int(gross_tot),
            "Net": int(net_tot),
            "Forever": int(forever_active),
            "Net AR %": net_ar,
            "Moltipl.": f"x{moltiplicatore}",
            "Compenso base reload": round(compenso_base, 2),
            "Compenso Forever Basic": round(compenso_forever_basic, 2),
            "Compenso Forever Premium": round(compenso_forever_premium, 2),
            "Compenso Plus": round(compenso_plus, 2),
            "Compenso EU": round(compenso_eu, 2),
            "Compenso Exchange": round(compenso_exchange, 2),
            "Subtotale (€)": round(compenso_subtotal, 2),
            "TOTALE stimato (€)": round(compenso_totale, 2),
            "_dettaglio_fasce": dettaglio_fasce,
            "_canale_int": canale_dom,  # per next-milestone
            "_net_ar_raw": net_ar,
            "_subtotal_raw": compenso_subtotal,
        })

    out = pd.DataFrame(rows).sort_values("TOTALE stimato (€)", ascending=False).reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────
#  FASE 3: EXTRA GARA RELOAD FOREVER TRIMESTRALE
# ─────────────────────────────────────────────────────────────────
# PPTX slide 17/33: gara per Ragione Sociale, target 100/200/300 Forever andate
# a buon fine per PDV. Premi 500€ / 1500€ / 3000€ per PDV.
# I target sono moltiplicati per il numero di PDV della RS.
# Periodo trimestrale (febbraio-aprile 2026 per l'edizione corrente).

EXTRA_FOREVER_SOGLIE_DEFAULT = [
    {"target_per_pdv": 100, "premio_per_pdv": 500},
    {"target_per_pdv": 200, "premio_per_pdv": 1500},
    {"target_per_pdv": 300, "premio_per_pdv": 3000},
]


def compute_extra_forever_trimestrale(df, soglie=None, forever_cumulato_manuale=None):
    """
    Calcola il compenso Extra Gara Reload Forever (gara trimestrale per RS).

    Per ogni RS:
    - Calcola le Forever andate a buon fine (R4_BASIC_NET + R4_PREMIUM_NET).
      Se non disponibili, usa Activeforever come stima.
    - Moltiplica i target per il numero di PDV della RS
    - Trova lo scaglione raggiunto (max premio applicabile)
    - Calcola anche il "gap" per il prossimo scaglione

    Parametri:
        forever_cumulato_manuale: dict {COMPANY_NAME: forever_cumulati_trimestre}
                                  per override manuale (utile se hai dati cumulati a parte)
    """
    if df.empty:
        return pd.DataFrame()

    soglie = soglie or EXTRA_FOREVER_SOGLIE_DEFAULT
    soglie_sorted = sorted(soglie, key=lambda x: x["target_per_pdv"])
    override = forever_cumulato_manuale or {}

    rows = []
    for company, sub in df.groupby("COMPANY_NAME", dropna=False):
        n_pdv = len(sub)
        company_str = safe_str(company)

        # Forever andate a buon fine: preferisco R4 Basic+Premium NET sales se disponibili
        r4_basic_net = float(sub.get("R4_BASIC_NET_SALES", pd.Series([0])).sum() or 0)
        r4_premium_net = float(sub.get("R4_PREMIUM_NET_SALES", pd.Series([0])).sum() or 0)
        forever_mese = r4_basic_net + r4_premium_net

        # Se l'utente ha fornito un cumulato trimestrale manuale, usa quello
        forever_periodo = override.get(company_str, forever_mese)

        # Target per la RS = target_per_pdv × n_pdv
        scaglione_raggiunto = 0
        premio = 0
        target_corrente = 0
        target_prossimo = None
        premio_prossimo = None
        for i, s in enumerate(soglie_sorted, 1):
            target_rs = s["target_per_pdv"] * n_pdv
            if forever_periodo >= target_rs:
                scaglione_raggiunto = i
                premio = s["premio_per_pdv"] * n_pdv
                target_corrente = target_rs
            else:
                target_prossimo = target_rs
                premio_prossimo = s["premio_per_pdv"] * n_pdv
                break

        gap_prossimo = (target_prossimo - forever_periodo) if target_prossimo is not None else None

        rows.append({
            "Ragione Sociale": company_str,
            "N PDV": n_pdv,
            "Forever periodo": int(forever_periodo),
            "Scaglione": scaglione_raggiunto,
            "Target raggiunto": int(target_corrente) if target_corrente else "—",
            "Premio (€)": int(premio),
            "Prossimo target": int(target_prossimo) if target_prossimo is not None else "MAX",
            "Gap": int(gap_prossimo) if gap_prossimo is not None else 0,
            "Premio prossimo (€)": int(premio_prossimo) if premio_prossimo is not None else 0,
            "_forever_periodo_raw": forever_periodo,
        })

    out = pd.DataFrame(rows).sort_values("Premio (€)", ascending=False).reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────
#  FASE 5: NEXT MILESTONE — quanto manca per passare allo scaglione successivo
# ─────────────────────────────────────────────────────────────────
def compute_next_milestone(compensi_df, soglie_ar=None):
    """
    Per ogni RS al di sotto del moltiplicatore X4, calcola quanti reload (NET)
    aggiuntivi servono per passare allo scaglione successivo (X1→X2, X2→X3, X3→X4).

    Algoritmo:
    - Per il moltiplicatore corrente, conosco la soglia AR target del prossimo
    - L'AR è (Net+PlusNet+Forever) / (TAM+Forever)
    - Per stimare i reload mancanti, calcolo il delta Net assumendo che TAM e Forever
      restino invariati: nuovo_AR = (Net + delta + Forever) / (TAM + Forever)
    - Risolvo per delta: delta = soglia/100 × (TAM+Forever) - (Net+PlusNet+Forever)
    - Calcolo anche il guadagno extra in € se passa allo scaglione (subtotale × differenza moltiplicatori)
    """
    if compensi_df.empty:
        return pd.DataFrame()

    soglie_ar = soglie_ar or SOGLIE_AR_DEFAULT
    rows = []

    # Recupero i raw values dai compensi
    for _, r in compensi_df.iterrows():
        canale = r.get("_canale_int", r.get("Canale", "Dealer"))
        net_ar = float(r.get("_net_ar_raw", r.get("Net AR %", 0)) or 0)
        subtot = float(r.get("_subtotal_raw", r.get("Subtotale (€)", 0)) or 0)
        molt_str = str(r.get("Moltipl.", "x1"))
        try:
            molt_curr = int(molt_str.replace("x", ""))
        except Exception:
            molt_curr = 1

        if molt_curr >= 4:
            continue  # già al massimo

        # Trova la soglia del prossimo scaglione
        s = soglie_ar.get(canale, soglie_ar.get("Dealer"))
        if molt_curr == 1:
            soglia_target = s["x2"]
            molt_next = 2
        elif molt_curr == 2:
            soglia_target = s["x3"]
            molt_next = 3
        else:  # molt_curr == 3
            soglia_target = s["x4"]
            molt_next = 4

        # Recupero TAM e Net dalla riga
        tam = float(r.get("TAM", 0) or 0)
        net = float(r.get("Net", 0) or 0)
        forever = float(r.get("Forever", 0) or 0)

        # delta_net necessario
        target_numeratore = soglia_target / 100.0 * (tam + forever)
        attuale_numeratore = (net_ar / 100.0) * (tam + forever) if (tam + forever) > 0 else 0
        delta_net = max(0, round(target_numeratore - attuale_numeratore, 1))

        # Se delta_net è enorme rispetto a Net attuale, è poco realistico → marca
        ratio = (delta_net / net) if net > 0 else float("inf")
        realistico = ratio <= 0.30  # entro +30% del Net attuale è realistico

        # Guadagno extra stimato
        guadagno_extra = subtot * (molt_next - molt_curr)

        rows.append({
            "Ragione Sociale": r["Ragione Sociale"],
            "Canale": canale,
            "Molt. attuale": f"x{molt_curr}",
            "Net AR %": net_ar,
            "Soglia prossima": f"{soglia_target}% → x{molt_next}",
            "Reload mancanti": int(delta_net),
            "Net attuale": int(net),
            "Realistico": "✅" if realistico else "—",
            "Guadagno extra (€)": round(guadagno_extra, 2),
            "_ratio": ratio,
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    # Ordino: prima i realistici per guadagno extra
    out = out.sort_values(
        ["Realistico", "Guadagno extra (€)"], ascending=[False, False]
    ).reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────
#  EXCEL / PDF EXPORT (con soglie configurabili — bug fix)
# ─────────────────────────────────────────────────────────────────
def _color_for_pct(val, soglia_critica, soglia_attenzione, RED, ORANGE, GREEN):
    try:
        v = float(val)
    except Exception:
        return None
    if v < soglia_critica:
        return RED
    if v < soglia_attenzione:
        return ORANGE
    return GREEN


def build_excel_riepilogo(kpis, meta, soglia_gross, soglia_net, sa_gross, sa_net, zona_label, regioni):
    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"
    BLUE = PatternFill("solid", fgColor="003087")
    RED = PatternFill("solid", fgColor="FFCDD2")
    ORANGE = PatternFill("solid", fgColor="FFE0B2")
    GREEN = PatternFill("solid", fgColor="C8E6C9")
    BD = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    C = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:O1")
    tc = ws["A1"]
    reg_str = " & ".join(regioni) if regioni else "—"
    tc.value = f"AVANZAMENTO RELOAD – {meta['month']:02d}/{meta['year']} – {reg_str} ({zona_label})"
    tc.font = Font(bold=True, color="FFFFFF", size=12)
    tc.fill = BLUE
    tc.alignment = C
    ws.row_dimensions[1].height = 30
    headers = ["Regione", "Area Manager", "District Manager", "N Negozi", "TAM",
               "Gross Sales", "Gross %", "Gross Δ", "Net Sales", "Net %", "Net Δ",
               "Forever Active", "Total Stores", "Forever %", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="0050B3")
        c.alignment = C
        c.border = BD
    ws.row_dimensions[2].height = 22

    def ds(v):
        try:
            if v is None or pd.isna(v):
                return "—"
        except Exception:
            pass
        return f"+{v}%" if v > 0 else f"{v}%"

    for ri, (_, row) in enumerate(kpis.iterrows(), 3):
        vals = [row["REGION"], row["AREA_MANAGER"], row["DISTRICT_MANAGER"],
                int(row["N_Negozi"]), int(row["TAM"]),
                int(row["Gross_Sales"]), row["Gross_%"], ds(row.get("Gross_Δ")),
                int(row["Net_Sales"]), row["Net_%"], ds(row.get("Net_Δ")),
                int(row["Forever_Active"]), int(row["Total_Stores"]),
                row["Forever_%"], row["status"].upper()]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = BD
            c.alignment = C
            if ci == 7:  # Gross %
                fill = _color_for_pct(row["Gross_%"], soglia_gross, sa_gross, RED, ORANGE, GREEN)
                if fill:
                    c.fill = fill
            if ci == 10:  # Net %
                fill = _color_for_pct(row["Net_%"], soglia_net, sa_net, RED, ORANGE, GREEN)
                if fill:
                    c.fill = fill
        ws.row_dimensions[ri].height = 18

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 38)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_dm(row, df_raw, meta, soglia_gross, soglia_net, sa_gross, sa_net,
                   compensi_df=None, calc_mode="standard"):
    """
    Report Excel completo per un District Manager — versione "chiusura mese".

    Foglio 1 — RIEPILOGO: KPI del DM, trend vs mese precedente, alert, messaggio WhatsApp
    Foglio 2 — STORE: tutti gli store con TAM/Gross/Net/Forever/Plus/Exchange/Gross%/Net%, colorato
    Foglio 3 — COMPENSI RS: compensi stimati per ogni Ragione Sociale del DM (se disponibili)
    Foglio 4 — TOP 3 CRITICI: i 3 store peggiori per Gross%, con azione suggerita
    """
    wb = Workbook()
    BLUE = PatternFill("solid", fgColor="003087")
    HEADER = PatternFill("solid", fgColor="0050B3")
    RED = PatternFill("solid", fgColor="FFCDD2")
    GREEN = PatternFill("solid", fgColor="C8E6C9")
    ORANGE = PatternFill("solid", fgColor="FFE0B2")
    LIGHT = PatternFill("solid", fgColor="E3F2FD")
    BD = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    C = Alignment(horizontal="center", vertical="center")
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dm_name = row["DISTRICT_MANAGER"]
    mese_str = datetime(meta["year"], meta["month"], 1).strftime("%B %Y").capitalize()

    sub = df_raw[df_raw["DISTRICT_MANAGER"] == dm_name].copy()
    sub["Gross_%"] = sub.apply(lambda r: safe_pct(r["SR_PLUS_GROSS_SALES"], r["TAM"]), axis=1)
    sub["Net_%"] = sub.apply(lambda r: safe_pct(r["SR_PLUS_NET_SALES"], r["TAM"]), axis=1)

    # ── FOGLIO 1: RIEPILOGO ──
    ws1 = wb.active
    ws1.title = "Riepilogo"
    # Header
    ws1.merge_cells("A1:H1")
    tc = ws1["A1"]
    tc.value = f"REPORT CHIUSURA MESE — {dm_name}"
    tc.font = Font(bold=True, color="FFFFFF", size=13)
    tc.fill = BLUE; tc.alignment = C
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:H2")
    ws1["A2"].value = f"{mese_str} — {row.get('REGION', '')} — AM: {row.get('AREA_MANAGER', '')}"
    ws1["A2"].font = Font(color="003087", size=10); ws1["A2"].alignment = C
    ws1.row_dimensions[2].height = 22

    # KPI box
    kpi_labels = ["Negozi", "TAM", "Gross Sales", "Gross %", "Net Sales", "Net %", "Forever", "Forever %"]
    kpi_values = [
        int(row["N_Negozi"]), int(row["TAM"]), int(row["Gross_Sales"]),
        f"{row['Gross_%']}%", int(row["Net_Sales"]), f"{row['Net_%']}%",
        int(row["Forever_Active"]), f"{row['Forever_%']}%",
    ]
    for ci, (lab, val) in enumerate(zip(kpi_labels, kpi_values), 1):
        c = ws1.cell(row=4, column=ci, value=lab)
        c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD
        c = ws1.cell(row=5, column=ci, value=val)
        c.font = Font(bold=True, size=12); c.alignment = C; c.border = BD
        # Colora Gross% e Net%
        if ci == 4:
            fill = _color_for_pct(row["Gross_%"], soglia_gross, sa_gross, RED, ORANGE, GREEN)
            if fill: c.fill = fill
        if ci == 6:
            fill = _color_for_pct(row["Net_%"], soglia_net, sa_net, RED, ORANGE, GREEN)
            if fill: c.fill = fill
    ws1.row_dimensions[4].height = 20
    ws1.row_dimensions[5].height = 28

    # Trend
    def _ds(v):
        try:
            if v is None or pd.isna(v): return "—"
        except Exception: pass
        return f"+{v}%" if v > 0 else f"{v}%"

    ws1["A7"] = "TREND vs mese precedente"
    ws1["A7"].font = Font(bold=True, color="003087", size=10)
    ws1["A8"] = "Gross Δ"; ws1["B8"] = _ds(row.get("Gross_Δ"))
    ws1["C8"] = "Net Δ";   ws1["D8"] = _ds(row.get("Net_Δ"))
    for c in [ws1["A8"], ws1["C8"]]:
        c.font = Font(bold=True, color="003087")

    # Alert
    alerts = row.get("alerts", [])
    ws1["A10"] = "ALERT"
    ws1["A10"].font = Font(bold=True, color="C62828", size=10)
    if alerts:
        for ai, a in enumerate(alerts):
            ws1.cell(row=11 + ai, column=1, value=f"⚠️ {a}").font = Font(color="C62828")
    else:
        ws1["A11"] = "✅ Tutti i KPI nella norma"
        ws1["A11"].font = Font(color="2E7D32")

    # Top 3 critici
    worst = sub.nsmallest(3, "Gross_%")
    ws1.cell(row=15, column=1, value="TOP 3 STORE CRITICI").font = Font(bold=True, color="C62828", size=10)
    for wi, (_, ws_row) in enumerate(worst.iterrows()):
        r = 16 + wi
        ws1.cell(row=r, column=1, value=ws_row["STORE"]).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=ws_row["CITY"])
        ws1.cell(row=r, column=3, value=f"Gross: {ws_row['Gross_%']}%").font = Font(color="C62828")
        ws1.cell(row=r, column=4, value=f"Net: {ws_row['Net_%']}%")
        ws1.cell(row=r, column=5, value=f"TAM: {int(ws_row['TAM'])}")
        ws1.cell(row=r, column=6, value=f"Forever: {int(ws_row['Activeforever'])}")

    # Messaggio WhatsApp
    msg = generate_message(row, meta)
    ws1.cell(row=21, column=1, value="MESSAGGIO WHATSAPP").font = Font(bold=True, color="003087", size=10)
    ws1.merge_cells("A22:H30")
    msg_cell = ws1["A22"]
    msg_cell.value = msg
    msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
    msg_cell.font = Font(size=9)

    for col in ws1.columns:
        ml = max(len(str(c.value or "")[:40]) for c in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 30)

    # ── FOGLIO 2: STORE ──
    ws2 = wb.create_sheet("Store")
    ws2.merge_cells("A1:N1")
    tc2 = ws2["A1"]
    tc2.value = f"Dettaglio Store — {dm_name} — {mese_str}"
    tc2.font = Font(bold=True, color="FFFFFF", size=11); tc2.fill = BLUE; tc2.alignment = C
    ws2.row_dimensions[1].height = 28

    store_cols = ["STORE", "COMPANY_NAME", "CITY", "PROVINCE_CODE", "STORE_TYPE", "TAM",
                  "SR_PLUS_GROSS_SALES", "Gross_%", "SR_PLUS_NET_SALES", "Net_%",
                  "Activeforever", "TotalStore"]
    # Aggiungi colonne opzionali se presenti
    extra_store = []
    for ec in ["PLUS_GROSS_SALES", "EXCHANGE_SALES", "R4_BASIC_GROSS_SALES"]:
        if ec in sub.columns:
            store_cols.append(ec)
            extra_store.append(ec)

    hdrs2 = ["Store", "Rag. Sociale", "Città", "Prov", "Tipo", "TAM",
             "Gross", "Gross%", "Net", "Net%", "Forever", "Tot Store"]
    hdrs2 += [{"PLUS_GROSS_SALES": "Plus", "EXCHANGE_SALES": "Exchange",
               "R4_BASIC_GROSS_SALES": "R4 Basic"}.get(e, e) for e in extra_store]

    for ci, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD

    sub_sorted = sub.sort_values("Gross_%", ascending=True)
    for ri, (_, sr) in enumerate(sub_sorted[store_cols].iterrows(), 3):
        for ci, val in enumerate(sr.values, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = BD; c.alignment = C
            gross_idx = 8  # Gross%
            net_idx = 10   # Net%
            if ci == gross_idx and isinstance(val, (int, float)):
                fill = _color_for_pct(val, soglia_gross, sa_gross, RED, ORANGE, GREEN)
                if fill: c.fill = fill
            if ci == net_idx and isinstance(val, (int, float)):
                fill = _color_for_pct(val, soglia_net, sa_net, RED, ORANGE, GREEN)
                if fill: c.fill = fill
        ws2.row_dimensions[ri].height = 17

    for col in ws2.columns:
        ml = max(len(str(c.value or "")[:35]) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 35)
    ws2.freeze_panes = "A3"

    # ── FOGLIO 3: COMPENSI RS (se disponibili) ──
    if compensi_df is not None and not compensi_df.empty:
        # Filtra RS del DM
        dm_companies = sub["COMPANY_NAME"].unique().tolist()
        comp_dm = compensi_df[compensi_df["Ragione Sociale"].isin([safe_str(c) for c in dm_companies])].copy()
        if not comp_dm.empty:
            ws3 = wb.create_sheet("Compensi RS")
            ws3.merge_cells("A1:J1")
            tc3 = ws3["A1"]
            tc3.value = f"Compensi Reload stimati — RS di {dm_name}"
            tc3.font = Font(bold=True, color="FFFFFF", size=11); tc3.fill = BLUE; tc3.alignment = C
            ws3.row_dimensions[1].height = 28

            comp_cols = ["Ragione Sociale", "Canale", "N PDV", "TAM", "Gross", "Net",
                         "Forever", "Net AR %", "Moltipl.", "TOTALE stimato (€)"]
            for ci, h in enumerate(comp_cols, 1):
                c = ws3.cell(row=2, column=ci, value=h)
                c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD

            for ri, (_, cr) in enumerate(comp_dm[comp_cols].iterrows(), 3):
                for ci, val in enumerate(cr.values, 1):
                    c = ws3.cell(row=ri, column=ci, value=val)
                    c.border = BD; c.alignment = C
                ws3.row_dimensions[ri].height = 17

            # Totale
            tot_row = ri + 1
            ws3.cell(row=tot_row, column=1, value="TOTALE DM").font = Font(bold=True, color="003087")
            ws3.cell(row=tot_row, column=10, value=round(comp_dm["TOTALE stimato (€)"].sum(), 2))
            ws3.cell(row=tot_row, column=10).font = Font(bold=True, size=12, color="003087")
            ws3.cell(row=tot_row, column=10).border = BD

            for col in ws3.columns:
                ml = max(len(str(c.value or "")[:35]) for c in col)
                ws3.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 35)
            ws3.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_am(am_name, kpis, df_raw, meta, soglia_gross, soglia_net, sa_gross, sa_net,
                   compensi_df=None, zona_label="A6"):
    """
    Report Excel per un Area Manager — versione "chiusura mese".

    Foglio 1 — RIEPILOGO AM: KPI aggregati dell'AM
    Foglio 2 — CLASSIFICA DM: tutti i DM dell'AM ordinati per Gross%, con status e trend
    Foglio 3 — TUTTI GLI STORE: store dell'AM in un'unica tabella
    """
    wb = Workbook()
    BLUE = PatternFill("solid", fgColor="003087")
    HEADER = PatternFill("solid", fgColor="0050B3")
    RED = PatternFill("solid", fgColor="FFCDD2")
    GREEN = PatternFill("solid", fgColor="C8E6C9")
    ORANGE = PatternFill("solid", fgColor="FFE0B2")
    BD = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    C = Alignment(horizontal="center", vertical="center")

    mese_str = datetime(meta["year"], meta["month"], 1).strftime("%B %Y").capitalize()
    am_kpis = kpis[kpis["AREA_MANAGER"] == am_name]
    am_stores = df_raw[df_raw["AREA_MANAGER"] == am_name].copy()

    # ── FOGLIO 1: RIEPILOGO AM ──
    ws1 = wb.active
    ws1.title = "Riepilogo AM"
    ws1.merge_cells("A1:H1")
    tc = ws1["A1"]
    tc.value = f"REPORT AREA MANAGER — {am_name}"
    tc.font = Font(bold=True, color="FFFFFF", size=13); tc.fill = BLUE; tc.alignment = C
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:H2")
    ws1["A2"].value = f"{mese_str} — Ambassador {zona_label}"
    ws1["A2"].font = Font(color="003087", size=10); ws1["A2"].alignment = C

    # KPI aggregati AM
    tam_am = int(am_kpis["TAM"].sum())
    gross_am = int(am_kpis["Gross_Sales"].sum())
    net_am = int(am_kpis["Net_Sales"].sum())
    forever_am = int(am_kpis["Forever_Active"].sum())
    n_dm = len(am_kpis)
    n_stores = int(am_kpis["N_Negozi"].sum())
    gross_pct_am = safe_pct(gross_am, tam_am)
    net_pct_am = safe_pct(net_am, tam_am)
    n_crit = (am_kpis["status"] == "critico").sum()
    n_ok = (am_kpis["status"] == "ok").sum()

    kpi_labels = ["N DM", "N Store", "TAM", "Gross", "Gross %", "Net", "Net %", "Forever"]
    kpi_values = [n_dm, n_stores, tam_am, gross_am, f"{gross_pct_am}%", net_am, f"{net_pct_am}%", forever_am]
    for ci, (lab, val) in enumerate(zip(kpi_labels, kpi_values), 1):
        c = ws1.cell(row=4, column=ci, value=lab)
        c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD
        c = ws1.cell(row=5, column=ci, value=val)
        c.font = Font(bold=True, size=12); c.alignment = C; c.border = BD
        if ci == 5:
            fill = _color_for_pct(gross_pct_am, soglia_gross, sa_gross, RED, ORANGE, GREEN)
            if fill: c.fill = fill
        if ci == 7:
            fill = _color_for_pct(net_pct_am, soglia_net, sa_net, RED, ORANGE, GREEN)
            if fill: c.fill = fill

    ws1["A7"] = f"Status: 🔴 Critici {n_crit} · ✅ OK {n_ok}"
    ws1["A7"].font = Font(bold=True, color="003087", size=10)

    for col in ws1.columns:
        ml = max(len(str(c.value or "")[:35]) for c in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 30)

    # ── FOGLIO 2: CLASSIFICA DM ──
    ws2 = wb.create_sheet("Classifica DM")
    ws2.merge_cells("A1:K1")
    tc2 = ws2["A1"]
    tc2.value = f"DM di {am_name} — ordinati per Gross% — {mese_str}"
    tc2.font = Font(bold=True, color="FFFFFF", size=11); tc2.fill = BLUE; tc2.alignment = C
    ws2.row_dimensions[1].height = 28

    dm_hdrs = ["Pos.", "District Manager", "Regione", "N Negozi", "TAM",
               "Gross", "Gross %", "Net", "Net %", "Forever %", "Status"]
    for ci, h in enumerate(dm_hdrs, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD

    am_sorted = am_kpis.sort_values("Gross_%", ascending=False)
    for pos, (_, dr) in enumerate(am_sorted.iterrows(), 1):
        r = pos + 2
        vals = [pos, dr["DISTRICT_MANAGER"], dr["REGION"], int(dr["N_Negozi"]),
                int(dr["TAM"]), int(dr["Gross_Sales"]), dr["Gross_%"],
                int(dr["Net_Sales"]), dr["Net_%"], dr["Forever_%"],
                dr["status"].upper()]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=r, column=ci, value=val)
            c.border = BD; c.alignment = C
            if ci == 7:
                fill = _color_for_pct(val, soglia_gross, sa_gross, RED, ORANGE, GREEN)
                if fill: c.fill = fill
            if ci == 9:
                fill = _color_for_pct(val, soglia_net, sa_net, RED, ORANGE, GREEN)
                if fill: c.fill = fill
            if ci == 11:
                if val == "CRITICO": c.font = Font(bold=True, color="C62828")
                elif val == "OK": c.font = Font(bold=True, color="2E7D32")
        ws2.row_dimensions[r].height = 18

    for col in ws2.columns:
        ml = max(len(str(c.value or "")[:35]) for c in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 35)
    ws2.freeze_panes = "A3"

    # ── FOGLIO 3: TUTTI GLI STORE ──
    ws3 = wb.create_sheet("Tutti Store")
    ws3.merge_cells("A1:L1")
    tc3 = ws3["A1"]
    tc3.value = f"Store di {am_name} — {mese_str}"
    tc3.font = Font(bold=True, color="FFFFFF", size=11); tc3.fill = BLUE; tc3.alignment = C
    ws3.row_dimensions[1].height = 28

    am_stores["Gross_%"] = am_stores.apply(lambda r: safe_pct(r["SR_PLUS_GROSS_SALES"], r["TAM"]), axis=1)
    am_stores["Net_%"] = am_stores.apply(lambda r: safe_pct(r["SR_PLUS_NET_SALES"], r["TAM"]), axis=1)

    st_cols = ["DISTRICT_MANAGER", "STORE", "COMPANY_NAME", "CITY", "PROVINCE_CODE",
               "STORE_TYPE", "TAM", "SR_PLUS_GROSS_SALES", "Gross_%",
               "SR_PLUS_NET_SALES", "Net_%", "Activeforever"]
    st_hdrs = ["DM", "Store", "Rag. Sociale", "Città", "Prov", "Tipo",
               "TAM", "Gross", "Gross%", "Net", "Net%", "Forever"]
    for ci, h in enumerate(st_hdrs, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = HEADER; c.alignment = C; c.border = BD

    am_sorted_stores = am_stores.sort_values(["DISTRICT_MANAGER", "Gross_%"], ascending=[True, True])
    for ri, (_, sr) in enumerate(am_sorted_stores[st_cols].iterrows(), 3):
        for ci, val in enumerate(sr.values, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.border = BD; c.alignment = C
            if ci == 9 and isinstance(val, (int, float)):
                fill = _color_for_pct(val, soglia_gross, sa_gross, RED, ORANGE, GREEN)
                if fill: c.fill = fill
            if ci == 11 and isinstance(val, (int, float)):
                fill = _color_for_pct(val, soglia_net, sa_net, RED, ORANGE, GREEN)
                if fill: c.fill = fill
        ws3.row_dimensions[ri].height = 17

    for col in ws3.columns:
        ml = max(len(str(c.value or "")[:35]) for c in col)
        ws3.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 35)
    ws3.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf_riepilogo(kpis, meta, soglia_gross, soglia_net, zona_label, regioni):
    mese_str = datetime(meta["year"], meta["month"], 1).strftime("%B %Y").upper()
    pdf = FPDF()
    pdf.add_page(orientation="L")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_fill_color(0, 48, 135)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 14)
    reg_str = " & ".join(regioni) if regioni else "—"
    pdf.cell(0, 12, latin1(f"WIND3 RELOAD – {mese_str} – {reg_str} ({zona_label})"),
             fill=True, ln=True, align="C")
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6,
             latin1(f"Soglie critiche: Gross < {soglia_gross}%  |  Net < {soglia_net}%  |  Forever = 0 -> CRITICO"),
             ln=True, align="C")
    pdf.ln(4)
    headers = ["District Manager", "Regione", "N.Neg", "TAM", "Gross", "Gross%",
               "Net", "Net%", "Forever%", "Trend G", "Trend N", "Status"]
    widths = [52, 26, 12, 16, 16, 18, 16, 18, 20, 18, 18, 20]
    pdf.set_fill_color(0, 80, 179)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(headers, widths):
        pdf.cell(w, 8, latin1(h), border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7.5)
    for _, row in kpis.iterrows():
        if row["status"] == "critico":
            pdf.set_fill_color(255, 205, 210)
        elif row["status"] == "attenzione":
            pdf.set_fill_color(255, 224, 178)
        else:
            pdf.set_fill_color(240, 248, 255)
        pdf.set_text_color(30, 30, 30)

        def safe_delta(v):
            try:
                if v is None or pd.isna(v):
                    return "—"
            except Exception:
                pass
            return f"+{v}%" if v > 0 else f"{v}%"

        vals = [
            safe_str(row["DISTRICT_MANAGER"])[:28],
            safe_str(row["REGION"])[:14],
            str(int(row["N_Negozi"])),
            str(int(row["TAM"])),
            str(int(row["Gross_Sales"])),
            f"{row['Gross_%']}%",
            str(int(row["Net_Sales"])),
            f"{row['Net_%']}%",
            f"{row['Forever_%']}%",
            safe_delta(row.get("Gross_Δ")),
            safe_delta(row.get("Net_Δ")),
            row["status"].upper(),
        ]
        for v, w in zip(vals, widths):
            pdf.cell(w, 7, latin1(str(v)), border=1, align="C", fill=True)
        pdf.ln()
    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, latin1(f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} – Wind3 Reload Dashboard {zona_label}"),
             align="C")
    return bytes(pdf.output())


# ─────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="w3-title" style="font-size:1.3rem">⚙️ Configurazione</div>', unsafe_allow_html=True)
    st.markdown("---")

    st.markdown('<div class="sidebar-section">🗺️ Zona</div>', unsafe_allow_html=True)
    zona_label = st.text_input("Ambassador code", "A6", help="Es: A1, A2, A6...")
    regioni_input = st.text_input("Regioni (separate da virgola)", "Campania, Puglia")
    regioni = [r.strip() for r in regioni_input.split(",") if r.strip()]
    filter_mode = st.selectbox(
        "Modalità filtro zona",
        ["AND", "OR", "AMB", "REG"],
        index=0,
        help=(
            "AND = ambassador E regione (consigliato)  •  "
            "OR = ambassador O regione  •  "
            "AMB = solo ambassador  •  "
            "REG = solo regioni"
        ),
    )

    st.markdown('<div class="sidebar-section">🎯 Soglie Critiche</div>', unsafe_allow_html=True)
    st.caption("Sotto queste soglie → 🔴 CRITICO")
    soglia_gross = st.slider("Gross % critica", 0, 100, 40, 1)
    soglia_net = st.slider("Net % critica", 0, 100, 37, 1)
    forever_zero_critico = st.checkbox("Forever = 0 → critico", value=True)

    st.markdown('<div class="sidebar-section">⚠️ Soglie Attenzione</div>', unsafe_allow_html=True)
    st.caption("Sotto queste soglie → 🟠 ATTENZIONE (sopra critica)")
    sa_gross = st.slider("Gross % attenzione", 0, 100, 60, 1)
    sa_net = st.slider("Net % attenzione", 0, 100, 50, 1)
    if sa_gross < soglia_gross:
        sa_gross = soglia_gross
    if sa_net < soglia_net:
        sa_net = soglia_net

    st.markdown('<div class="sidebar-section">🔄 Modalità calcolo</div>', unsafe_allow_html=True)
    st.caption("Includere le Forever nel calcolo Gross/Net?")
    calc_mode_label = st.radio(
        "Calcolo KPI",
        ["Standard", "Adjusted (con Forever)", "Confronto (entrambi)"],
        index=1,
        help=(
            "Standard: Gross/TAM, Net/TAM (calcolo classico Wind3)\n"
            "Adjusted: (Gross+Forever)/(TAM+Forever) — realtà del lavoro\n"
            "Confronto: mostra entrambi affiancati"
        ),
        label_visibility="collapsed",
    )
    calc_mode = {"Standard": "standard",
                 "Adjusted (con Forever)": "adjusted",
                 "Confronto (entrambi)": "confronto"}[calc_mode_label]

    st.markdown('<div class="sidebar-section">📑 Foglio Excel</div>', unsafe_allow_html=True)
    sheet_name = st.text_input("Nome foglio", "Sales x Store")

    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:0.78rem;color:#7B97BF;line-height:1.8">
    🔴 Gross critico: <b style="color:#EF5350">< {soglia_gross}%</b><br>
    🟠 Gross attenzione: <b style="color:#FFA726">< {sa_gross}%</b><br>
    🔴 Net critico: <b style="color:#EF5350">< {soglia_net}%</b><br>
    🟠 Net attenzione: <b style="color:#FFA726">< {sa_net}%</b><br>
    🔄 Modalità: <b style="color:#00A8E0">{calc_mode_label}</b><br>
    🗺️ Zona: <b style="color:#00A8E0">{zona_label}</b> – {', '.join(regioni)}<br>
    ⚙️ Filtro: <b>{filter_mode}</b>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="w3-header">
    <div class="w3-glow"></div>
    <div class="w3-title">WIND<span>3</span> RELOAD</div>
    <div class="w3-sub">Sales Performance Dashboard · v4</div>
    <div class="w3-badge">📶 Ambassador {zona_label}</div>
    <div class="w3-badge">🗺️ {' & '.join(regioni) if regioni else '—'}</div>
    <div class="w3-badge">⚙️ Filtro {filter_mode}</div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  UPLOAD
# ─────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("📎 Carica il file Excel di avanzamento", type=["xlsx"])

if uploaded is None:
    st.markdown("""
    <div style="text-align:center;padding:4rem 2rem;color:#7B97BF;">
        <div style="font-size:3rem;margin-bottom:1rem;">📂</div>
        <div style="font-family:'Barlow Condensed';font-size:1.3rem;font-weight:700;margin-bottom:0.5rem;">
            Carica il file Excel per iniziare
        </div>
        <div style="font-size:0.85rem;">Formato: <code>Avanzamento_DD-MM-YY__Sales_x_Store.xlsx</code></div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────────────────────────
#  ANALISI
# ─────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_df(file_bytes, amb, reg_str, sheet_name, filter_mode):
    cfg = {
        "ambassador_filter": [amb] if amb else [],
        "regioni_filter": [r.strip() for r in reg_str.split(",") if r.strip()],
        "sheet_name": sheet_name,
        "filter_mode": filter_mode,
    }
    df, df_prev, meta = load_and_filter(io.BytesIO(file_bytes), cfg)
    return df, df_prev, meta


file_bytes = uploaded.read()

with st.spinner("Analisi in corso..."):
    try:
        df, df_prev, meta = load_df(file_bytes, zona_label, ",".join(regioni), sheet_name, filter_mode)
    except Exception as e:
        st.error(f"Errore nel caricamento: {e}")
        st.stop()

# KPI calcolati FUORI dalla cache (così le soglie cambiano subito)
cfg_kpi = {
    "soglia_critica_gross": soglia_gross,
    "soglia_critica_net": soglia_net,
    "soglia_attenzione_gross": sa_gross,
    "soglia_attenzione_net": sa_net,
    "forever_zero_critico": forever_zero_critico,
    "calc_mode": calc_mode,
}
kpis = compute_kpis(df, df_prev, cfg_kpi)

if kpis.empty:
    st.warning("Nessun KPI calcolabile sui dati filtrati.")
    st.stop()

# Aggregati per Ragione Sociale (Fase 1)
rs_df = compute_rs_aggregates(df, calc_mode=calc_mode)

# Compensi reload (Fase 2 + Fase 4 raffinata)
compensi_df = compute_compensi_reload(df)

# Extra Gara Reload Forever trimestrale (Fase 3)
extra_forever_df = compute_extra_forever_trimestrale(df)

# Next milestone — quanto manca per passare allo scaglione successivo (Fase 5)
next_milestone_df = compute_next_milestone(compensi_df)

if kpis.empty:
    st.warning("Nessun KPI calcolabile sui dati filtrati.")
    st.stop()

mese_str = datetime(meta["year"], meta["month"], 1).strftime("%B %Y").capitalize()
n_critici = (kpis["status"] == "critico").sum()
n_attenzione = (kpis["status"] == "attenzione").sum()
n_ok = (kpis["status"] == "ok").sum()
avg_gross = kpis["Gross_%"].mean()
avg_net = kpis["Net_%"].mean()

if st.session_state["selected_dm"] is None or st.session_state["selected_dm"] not in kpis["DISTRICT_MANAGER"].values:
    st.session_state["selected_dm"] = kpis["DISTRICT_MANAGER"].iloc[0]


# ─────────────────────────────────────────────────────────────────
#  KPI TILES
# ─────────────────────────────────────────────────────────────────
n_am = kpis["AREA_MANAGER"].nunique()
mode_badge = {"standard": "STD", "adjusted": "ADJ", "confronto": "STD/ADJ"}[calc_mode]
st.markdown(f"""
<div class="kpi-row">
    <div class="kpi-tile accent">
        <div class="kpi-label">📅 Periodo · {mode_badge}</div>
        <div class="kpi-value accent" style="font-size:1.3rem">{mese_str}</div>
        <div class="kpi-sub">{len(df)} store · {len(kpis)} DM · {n_am} AM · {len(rs_df)} RS</div>
    </div>
    <div class="kpi-tile {'red' if avg_gross<soglia_gross else 'green'}">
        <div class="kpi-label">📊 Gross medio</div>
        <div class="kpi-value {'red' if avg_gross<soglia_gross else 'green'}">{avg_gross:.1f}%</div>
        <div class="kpi-sub">Soglia: {soglia_gross}%</div>
    </div>
    <div class="kpi-tile {'red' if avg_net<soglia_net else 'green'}">
        <div class="kpi-label">📈 Net medio</div>
        <div class="kpi-value {'red' if avg_net<soglia_net else 'green'}">{avg_net:.1f}%</div>
        <div class="kpi-sub">Soglia: {soglia_net}%</div>
    </div>
    <div class="kpi-tile red">
        <div class="kpi-label">🔴 Critici</div>
        <div class="kpi-value red">{n_critici}</div>
        <div class="kpi-sub">su {len(kpis)} DM</div>
    </div>
    <div class="kpi-tile orange">
        <div class="kpi-label">⚠️ Attenzione</div>
        <div class="kpi-value" style="color:#FFA726">{n_attenzione}</div>
        <div class="kpi-sub">in fascia</div>
    </div>
    <div class="kpi-tile green">
        <div class="kpi-label">✅ OK</div>
        <div class="kpi-value green">{n_ok}</div>
        <div class="kpi-sub">sopra soglie</div>
    </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  TABS
# ─────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab_rs, tab_comp, tab4 = st.tabs([
    "📋 Riepilogo DM",
    "📄 Dettaglio & Messaggi",
    "🌳 Mappa Zona",
    "🏢 Ragioni Sociali",
    "💰 Compensi",
    "⬇️ Download",
])


# ── TAB 1: RIEPILOGO DM ────────────────────────────────────────
with tab1:
    col_f1, col_f2, col_f3, col_f4 = st.columns([1, 1, 1, 2])
    with col_f1:
        filtro_status = st.selectbox("Status", ["Tutti", "🔴 Critici", "⚠️ Attenzione", "✅ OK"])
    with col_f2:
        filtro_regione = st.selectbox("Regione", ["Tutte"] + sorted(kpis["REGION"].dropna().unique().tolist()))
    with col_f3:
        store_types = ["Tutti"] + sorted(df["STORE_TYPE"].dropna().unique().tolist())
        filtro_tipo = st.selectbox("Tipo Store", store_types)
    with col_f4:
        filtro_am = st.selectbox(
            "Area Manager",
            ["Tutti"] + sorted(kpis["AREA_MANAGER"].dropna().unique().tolist()),
        )

    col_s1, col_s2 = st.columns([2, 1])
    with col_s1:
        sort_col = st.selectbox(
            "Ordina per",
            ["Gross_%", "Net_%", "Forever_%", "N_Negozi", "TAM", "DISTRICT_MANAGER"],
            format_func=lambda x: {
                "Gross_%": "Gross %", "Net_%": "Net %", "Forever_%": "Forever %",
                "N_Negozi": "N. Negozi", "TAM": "TAM", "DISTRICT_MANAGER": "Nome",
            }[x],
        )
    with col_s2:
        sort_asc = st.checkbox("Crescente", value=False)

    filtered = kpis.copy()
    if filtro_status == "🔴 Critici":
        filtered = filtered[filtered["status"] == "critico"]
    elif filtro_status == "⚠️ Attenzione":
        filtered = filtered[filtered["status"] == "attenzione"]
    elif filtro_status == "✅ OK":
        filtered = filtered[filtered["status"] == "ok"]
    if filtro_regione != "Tutte":
        filtered = filtered[filtered["REGION"] == filtro_regione]
    if filtro_am != "Tutti":
        filtered = filtered[filtered["AREA_MANAGER"] == filtro_am]

    if filtro_tipo != "Tutti":
        df_tipo = df[df["STORE_TYPE"] == filtro_tipo]
        at = df_tipo.groupby("DISTRICT_MANAGER").agg(
            TAM_t=("TAM", "sum"),
            Gross_t=("SR_PLUS_GROSS_SALES", "sum"),
            Net_t=("SR_PLUS_NET_SALES", "sum"),
            N_t=("SHOP_CODE", "count"),
        ).reset_index()
        at["Gross_%_t"] = at.apply(lambda r: safe_pct(r["Gross_t"], r["TAM_t"]), axis=1)
        at["Net_%_t"] = at.apply(lambda r: safe_pct(r["Net_t"], r["TAM_t"]), axis=1)
        filtered = filtered.merge(
            at[["DISTRICT_MANAGER", "Gross_%_t", "Net_%_t", "N_t"]],
            on="DISTRICT_MANAGER", how="inner",
        )

    filtered = filtered.sort_values(sort_col, ascending=sort_asc)

    st.markdown(
        f'<div class="section-title">District Manager — {len(filtered)} risultati · {sort_col} {"↑" if sort_asc else "↓"}</div>',
        unsafe_allow_html=True,
    )

    show_chart = st.checkbox("📊 Mostra grafico Gross/Net/Forever", value=False)
    if show_chart and not filtered.empty:
        import altair as alt
        cd = filtered[["DISTRICT_MANAGER", "Gross_%", "Net_%", "Forever_%"]].rename(
            columns={"DISTRICT_MANAGER": "DM", "Gross_%": "Gross", "Net_%": "Net", "Forever_%": "Forever"})
        cm = cd.melt("DM", var_name="KPI", value_name="Valore")
        chart = alt.Chart(cm).mark_bar().encode(
            x=alt.X("DM:N", sort=None, axis=alt.Axis(labelAngle=-45, labelLimit=120)),
            y=alt.Y("Valore:Q", title="%"),
            color=alt.Color("KPI:N", scale=alt.Scale(domain=["Gross", "Net", "Forever"],
                                                     range=["#1976D2", "#00A8E0", "#4CAF50"])),
            xOffset="KPI:N",
            tooltip=["DM", "KPI", "Valore"],
        ).properties(height=300).configure_view(strokeOpacity=0).configure(
            background="transparent"
        ).configure_axis(labelColor="#7B97BF", titleColor="#7B97BF", gridColor="#0C1F35"
                         ).configure_legend(labelColor="#E8F0FE", titleColor="#7B97BF")
        st.altair_chart(chart, use_container_width=True)

    def ckpi(val, s, sa):
        if val < s:
            return "red"
        if val < sa:
            return "orange"
        return "green"

    def thtml(d):
        if d is None or pd.isna(d):
            return '<span class="dm-trend flat">—</span>'
        if d > 0:
            return f'<span class="dm-trend up">▲ +{d}%</span>'
        if d < 0:
            return f'<span class="dm-trend down">▼ {d}%</span>'
        return '<span class="dm-trend flat">= 0%</span>'

    for _idx, row in filtered.iterrows():
        dm_name = row["DISTRICT_MANAGER"]
        alerts_txt = " | ".join(row["alerts"]) if row["alerts"] else ""
        f_icon = "🔴" if row["Forever_Active"] == 0 else "♾️"
        g_show = row.get("Gross_%_t", row["Gross_%"])
        n_show = row.get("Net_%_t", row["Net_%"])
        tipo_note = f" ({filtro_tipo})" if filtro_tipo != "Tutti" else ""

        col_card, col_btn = st.columns([11, 1])
        with col_card:
            st.markdown(f"""
            <div class="dm-card {row['status']}">
                <div class="dm-name">{dm_name}<div class="dm-region">{row['REGION']} · {row['AREA_MANAGER']}</div></div>
                <div class="dm-kpi">
                    <div class="dm-kpi-label">Gross{tipo_note}</div>
                    <div class="dm-kpi-val {ckpi(g_show,soglia_gross,sa_gross)}">{g_show}%</div>
                    {thtml(row.get('Gross_Δ'))}
                </div>
                <div class="dm-kpi">
                    <div class="dm-kpi-label">Net{tipo_note}</div>
                    <div class="dm-kpi-val {ckpi(n_show,soglia_net,sa_net)}">{n_show}%</div>
                    {thtml(row.get('Net_Δ'))}
                </div>
                <div class="dm-forever">
                    {f_icon} <strong>{int(row['Forever_Active'])}</strong>/{int(row['Total_Stores'])}<br>
                    <span style="font-size:0.8rem">{row['Forever_%']}%</span>
                </div>
                <div class="dm-alerts">{alerts_txt}</div>
            </div>""", unsafe_allow_html=True)
        with col_btn:
            if st.button("🏪", key=f"btn_{_idx}_{dm_name}", help=f"Vedi store di {dm_name}"):
                st.session_state["selected_dm"] = dm_name
                st.session_state["goto_detail"] = True
                st.rerun()

    if st.session_state.get("goto_detail"):
        st.info("👆 Vai al tab **📄 Dettaglio & Messaggi** per vedere i negozi del DM selezionato")
        st.session_state["goto_detail"] = False


# ── TAB 2: DETTAGLIO & MESSAGGI ────────────────────────────────
with tab2:
    sub_tab1, sub_tab2 = st.tabs(["👤 Per District Manager", "🔍 Cerca Store / Ragione Sociale"])

    # Precomputa mappa stato per DM (fix performance format_func O(n²))
    status_by_dm = dict(zip(kpis["DISTRICT_MANAGER"], kpis["status"]))

    def dm_label(x):
        s = status_by_dm.get(x, "ok")
        icon = "🔴" if s == "critico" else ("⚠️" if s == "attenzione" else "✅")
        return f"{icon} {x}"

    with sub_tab1:
        dm_list = kpis["DISTRICT_MANAGER"].tolist()
        saved_dm = st.session_state.get("selected_dm", dm_list[0])
        dm_idx = dm_list.index(saved_dm) if saved_dm in dm_list else 0

        dm_selected = st.selectbox(
            "Seleziona District Manager", dm_list, index=dm_idx, format_func=dm_label,
        )
        st.session_state["selected_dm"] = dm_selected
        row_sel = kpis[kpis["DISTRICT_MANAGER"] == dm_selected].iloc[0]

        col_msg, col_store = st.columns([1, 2])

        with col_msg:
            st.markdown('<div class="section-title">💬 Messaggio pronto</div>', unsafe_allow_html=True)
            msg = generate_message(row_sel, meta)
            st.text_area("msg", msg, height=320, label_visibility="collapsed")
            st.download_button(
                "⬇️ Scarica .txt", data=msg.encode("utf-8"),
                file_name=f"{dm_selected.replace(' ','_')}_{meta['month']:02d}_{meta['year']}.txt",
                mime="text/plain", key="dl_msg_dm",
            )

        with col_store:
            st.markdown('<div class="section-title">🏪 Dettaglio Store</div>', unsafe_allow_html=True)
            sub = df[df["DISTRICT_MANAGER"] == dm_selected].copy()
            sub["Gross_%"] = sub.apply(lambda r: safe_pct(r["SR_PLUS_GROSS_SALES"], r["TAM"]), axis=1)
            sub["Net_%"] = sub.apply(lambda r: safe_pct(r["SR_PLUS_NET_SALES"], r["TAM"]), axis=1)

            tipi_dm = ["Tutti"] + sorted(sub["STORE_TYPE"].dropna().unique().tolist())
            ft_dm = st.selectbox("Filtra per tipo store", tipi_dm, key="tipo_dm")
            if ft_dm != "Tutti":
                sub = sub[sub["STORE_TYPE"] == ft_dm]

            display_cols = {
                "STORE": "Store", "COMPANY_NAME": "Ragione Sociale", "CITY": "Città",
                "PROVINCE_CODE": "Prov", "STORE_TYPE": "Tipo", "TAM": "TAM",
                "SR_PLUS_GROSS_SALES": "Gross", "Gross_%": "Gross%",
                "SR_PLUS_NET_SALES": "Net", "Net_%": "Net%", "Activeforever": "Forever",
            }
            sub_display = sub[list(display_cols.keys())].rename(columns=display_cols)

            sc_store = st.selectbox("Ordina store per", ["Gross%", "Net%", "TAM", "Store", "Città"], key="sort_store")
            sa_store = st.checkbox("Crescente", value=True, key="asc_store")
            sub_display = sub_display.sort_values(sc_store, ascending=sa_store)

            def cg(v):
                if v < soglia_gross:
                    return "background-color:#FFCDD2;color:#B71C1C"
                if v < sa_gross:
                    return "background-color:#FFE0B2;color:#E65100"
                return "background-color:#C8E6C9;color:#1B5E20"

            def cn(v):
                if v < soglia_net:
                    return "background-color:#FFCDD2;color:#B71C1C"
                if v < sa_net:
                    return "background-color:#FFE0B2;color:#E65100"
                return "background-color:#C8E6C9;color:#1B5E20"

            styled = sub_display.style.map(cg, subset=["Gross%"]).map(cn, subset=["Net%"])
            st.dataframe(styled, use_container_width=True, height=350)

            buf_dm = build_excel_dm(row_sel, df, meta, soglia_gross, soglia_net, sa_gross, sa_net, compensi_df=compensi_df)
            st.download_button(
                "⬇️ Scarica Excel DM", data=buf_dm,
                file_name=f"{dm_selected.replace(' ','_')}_{meta['month']:02d}_{meta['year']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_dm",
            )

    with sub_tab2:
        st.markdown('<div class="section-title">🔍 Ricerca Store / Ragione Sociale</div>', unsafe_allow_html=True)
        col_sq, col_st1, col_st2 = st.columns([3, 1, 1])
        with col_sq:
            search_query = st.text_input(
                "Cerca", placeholder="Es: MEDIAWORLD, BLOCKCHAIN SRL, VIA ROMA, NAPOLI...",
                label_visibility="collapsed",
            )
        with col_st1:
            ft_search = st.selectbox(
                "Tipo", ["Tutti"] + sorted(df["STORE_TYPE"].dropna().unique().tolist()),
                key="tipo_search", label_visibility="collapsed",
            )
        with col_st2:
            ft_region_search = st.selectbox(
                "Regione", ["Tutte"] + sorted(df["REGION"].dropna().unique().tolist()),
                key="reg_search", label_visibility="collapsed",
            )

        if search_query.strip():
            q = search_query.strip().upper()
            search_cols = ["STORE", "COMPANY_NAME", "CITY", "STORE_ADDRESS", "PROVINCE_CODE"]
            mask = pd.Series(False, index=df.index)
            for c in search_cols:
                if c in df.columns:
                    mask |= df[c].astype(str).str.upper().str.contains(q, na=False)
            results = df[mask].copy()
            if ft_search != "Tutti":
                results = results[results["STORE_TYPE"] == ft_search]
            if ft_region_search != "Tutte":
                results = results[results["REGION"] == ft_region_search]

            if results.empty:
                st.info(f"Nessun risultato per: {search_query}")
            else:
                results["Gross_%"] = results.apply(lambda r: safe_pct(r["SR_PLUS_GROSS_SALES"], r["TAM"]), axis=1)
                results["Net_%"] = results.apply(lambda r: safe_pct(r["SR_PLUS_NET_SALES"], r["TAM"]), axis=1)
                st.markdown(f"**{len(results)} store trovati**")
                rc = {
                    "AREA_MANAGER": "AM", "DISTRICT_MANAGER": "District Manager",
                    "STORE": "Store", "COMPANY_NAME": "Ragione Sociale",
                    "CITY": "Città", "PROVINCE_CODE": "Prov", "STORE_TYPE": "Tipo", "TAM": "TAM",
                    "SR_PLUS_GROSS_SALES": "Gross", "Gross_%": "Gross%",
                    "SR_PLUS_NET_SALES": "Net", "Net_%": "Net%", "Activeforever": "Forever",
                }
                rd = results[list(rc.keys())].rename(columns=rc).sort_values("Gross%")

                def cg2(v):
                    if v < soglia_gross:
                        return "background-color:#FFCDD2;color:#B71C1C"
                    if v < sa_gross:
                        return "background-color:#FFE0B2;color:#E65100"
                    return "background-color:#C8E6C9;color:#1B5E20"

                def cn2(v):
                    if v < soglia_net:
                        return "background-color:#FFCDD2;color:#B71C1C"
                    if v < sa_net:
                        return "background-color:#FFE0B2;color:#E65100"
                    return "background-color:#C8E6C9;color:#1B5E20"

                st.dataframe(
                    rd.style.map(cg2, subset=["Gross%"]).map(cn2, subset=["Net%"]),
                    use_container_width=True, height=400,
                )
        else:
            st.markdown(f"""
            <div style="text-align:center;padding:2.5rem;color:#7B97BF;">
                <div style="font-size:2rem;margin-bottom:0.5rem;">🔍</div>
                Digita nome store, ragione sociale, città, indirizzo o provincia<br>
                tra i tuoi <b style="color:#00A8E0">{len(df)}</b> store
            </div>""", unsafe_allow_html=True)


# ── TAB 3: MAPPA ZONA (gerarchia) ──────────────────────────────
with tab3:
    st.markdown(
        '<div class="section-title">🌳 Mappa gerarchica della zona</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        f"Vista ad albero di tutti gli **Area Manager**, i **District Manager** e i "
        f"**negozi** che competono alla zona {zona_label} ({' & '.join(regioni) if regioni else '—'})."
    )

    col_h1, col_h2, col_h3 = st.columns([1, 1, 1])
    with col_h1:
        h_filter_am = st.selectbox(
            "Filtra AM", ["Tutti"] + sorted(df["AREA_MANAGER"].dropna().unique().tolist()),
            key="h_am",
        )
    with col_h2:
        h_filter_type = st.selectbox(
            "Filtra tipo store",
            ["Tutti"] + sorted(df["STORE_TYPE"].dropna().unique().tolist()),
            key="h_type",
        )
    with col_h3:
        h_show_stores = st.checkbox("Mostra elenco negozi", value=False, key="h_stores")

    df_h = df.copy()
    if h_filter_am != "Tutti":
        df_h = df_h[df_h["AREA_MANAGER"] == h_filter_am]
    if h_filter_type != "Tutti":
        df_h = df_h[df_h["STORE_TYPE"] == h_filter_type]

    tree = build_hierarchy(df_h)

    # Riassunto
    tot_am = len(tree)
    tot_dm = sum(len(v["dms"]) for v in tree.values())
    tot_stores = sum(v["n_stores"] for v in tree.values())
    st.markdown(f"""
    <div style="margin:0.6rem 0 1rem;font-size:0.85rem;color:#7B97BF">
        <b style="color:#00A8E0">{tot_am}</b> Area Manager ·
        <b style="color:#00A8E0">{tot_dm}</b> District Manager ·
        <b style="color:#00A8E0">{tot_stores}</b> negozi
    </div>
    """, unsafe_allow_html=True)

    # ── ANALISI GROSS & NET ───────────────────────────────────
    st.markdown('<div class="section-title">📊 Analisi Gross & Net</div>', unsafe_allow_html=True)
    analysis = analyze_gross_net(df_h, soglia_gross, soglia_net, sa_gross, sa_net)

    if analysis:
        # Riga metriche aggregate
        col_a1, col_a2, col_a3, col_a4 = st.columns(4)
        with col_a1:
            color_g = "red" if analysis["gross_pct"] < soglia_gross else ("orange" if analysis["gross_pct"] < sa_gross else "green")
            st.markdown(f"""
            <div class="kpi-tile {color_g}">
                <div class="kpi-label">🔵 Gross zona</div>
                <div class="kpi-value {color_g}">{analysis['gross_pct']}%</div>
                <div class="kpi-sub">{analysis['gross']:,} / {analysis['tam']:,}</div>
            </div>""", unsafe_allow_html=True)
        with col_a2:
            color_n = "red" if analysis["net_pct"] < soglia_net else ("orange" if analysis["net_pct"] < sa_net else "green")
            st.markdown(f"""
            <div class="kpi-tile {color_n}">
                <div class="kpi-label">🟢 Net zona</div>
                <div class="kpi-value {color_n}">{analysis['net_pct']}%</div>
                <div class="kpi-sub">{analysis['net']:,} / {analysis['tam']:,}</div>
            </div>""", unsafe_allow_html=True)
        with col_a3:
            st.markdown(f"""
            <div class="kpi-tile accent">
                <div class="kpi-label">🔄 Conversione Net/Gross</div>
                <div class="kpi-value accent">{analysis['net_on_gross']}%</div>
                <div class="kpi-sub">qualità delle vendite</div>
            </div>""", unsafe_allow_html=True)
        with col_a4:
            st.markdown(f"""
            <div class="kpi-tile accent">
                <div class="kpi-label">📍 Mediana store</div>
                <div class="kpi-value accent" style="font-size:1.4rem">G {analysis['median_gross_pct']:.1f}%<br>N {analysis['median_net_pct']:.1f}%</div>
                <div class="kpi-sub">50° percentile</div>
            </div>""", unsafe_allow_html=True)

        st.markdown(" ")
        # Gap necessario per raggiungere la soglia attenzione
        if analysis["needed_gross"] > 0 or analysis["needed_net"] > 0:
            gap_msgs = []
            if analysis["needed_gross"] > 0:
                gap_msgs.append(f"<b style='color:#FFA726'>+{analysis['needed_gross']:,}</b> Gross per superare il {sa_gross}%")
            if analysis["needed_net"] > 0:
                gap_msgs.append(f"<b style='color:#FFA726'>+{analysis['needed_net']:,}</b> Net per superare il {sa_net}%")
            st.markdown(f"""
            <div style="background:rgba(255,167,38,0.08);border-left:3px solid #FFA726;
                padding:0.7rem 1rem;border-radius:6px;font-size:0.85rem;margin:0.5rem 0;">
                ⚠️ <b>Gap rispetto alla soglia attenzione:</b><br>
                {' &nbsp;·&nbsp; '.join(gap_msgs)}
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:rgba(102,187,106,0.08);border-left:3px solid #66BB6A;
                padding:0.7rem 1rem;border-radius:6px;font-size:0.85rem;margin:0.5rem 0;">
                ✅ <b>Zona sopra le soglie di attenzione</b> sia su Gross che su Net
            </div>
            """, unsafe_allow_html=True)

        # Distribuzione store per fascia
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("**Distribuzione store · Gross**")
            for label, n in analysis["fasce_gross"].items():
                pct = safe_pct(n, analysis["n_stores"])
                bar_w = min(100, pct)
                st.markdown(f"""
                <div style="font-size:0.78rem;margin-bottom:0.3rem;">
                    <div style="display:flex;justify-content:space-between;color:#E8F0FE">
                        <span>{label}</span><span><b>{n}</b> ({pct}%)</span>
                    </div>
                    <div style="background:#0C1F35;height:6px;border-radius:3px;margin-top:2px;">
                        <div style="background:#00A8E0;width:{bar_w}%;height:100%;border-radius:3px;"></div>
                    </div>
                </div>""", unsafe_allow_html=True)
        with col_f2:
            st.markdown("**Distribuzione store · Net**")
            for label, n in analysis["fasce_net"].items():
                pct = safe_pct(n, analysis["n_stores"])
                bar_w = min(100, pct)
                st.markdown(f"""
                <div style="font-size:0.78rem;margin-bottom:0.3rem;">
                    <div style="display:flex;justify-content:space-between;color:#E8F0FE">
                        <span>{label}</span><span><b>{n}</b> ({pct}%)</span>
                    </div>
                    <div style="background:#0C1F35;height:6px;border-radius:3px;margin-top:2px;">
                        <div style="background:#1976D2;width:{bar_w}%;height:100%;border-radius:3px;"></div>
                    </div>
                </div>""", unsafe_allow_html=True)

        # Top / Bottom store
        st.markdown(" ")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            st.markdown("**🏆 Top 5 store · Gross%**")
            for s in analysis["top_gross"]:
                st.markdown(f"""
                <div style="font-size:0.78rem;padding:0.3rem 0;border-bottom:1px solid #0C1F35;">
                    <span style="color:#66BB6A;font-weight:700">{s['g_pct']}%</span> ·
                    <b style="color:#E8F0FE">{safe_str(s['STORE'])[:40]}</b><br>
                    <span style="color:#7B97BF;font-size:0.7rem">{safe_str(s['DISTRICT_MANAGER'])} · TAM {int(s['TAM'])}</span>
                </div>""", unsafe_allow_html=True)
        with col_t2:
            st.markdown("**⚠️ Bottom 5 store · Gross% (TAM>0)**")
            for s in analysis["bot_gross"]:
                st.markdown(f"""
                <div style="font-size:0.78rem;padding:0.3rem 0;border-bottom:1px solid #0C1F35;">
                    <span style="color:#EF5350;font-weight:700">{s['g_pct']}%</span> ·
                    <b style="color:#E8F0FE">{safe_str(s['STORE'])[:40]}</b><br>
                    <span style="color:#7B97BF;font-size:0.7rem">{safe_str(s['DISTRICT_MANAGER'])} · TAM {int(s['TAM'])}</span>
                </div>""", unsafe_allow_html=True)

        # Per tipo store
        if not analysis["by_type"].empty and len(analysis["by_type"]) > 1:
            st.markdown(" ")
            st.markdown("**📦 Performance per tipo store**")
            by_type_display = analysis["by_type"][["STORE_TYPE", "n", "tam", "gross", "gross_pct", "net", "net_pct", "share_gross"]].rename(
                columns={
                    "STORE_TYPE": "Tipo", "n": "N. store", "tam": "TAM",
                    "gross": "Gross", "gross_pct": "Gross %",
                    "net": "Net", "net_pct": "Net %", "share_gross": "Quota Gross %",
                }
            )
            st.dataframe(by_type_display, use_container_width=True, hide_index=True)

        if analysis["n_stores_zero_tam"] > 0:
            st.caption(f"ℹ️ {analysis['n_stores_zero_tam']} store con TAM=0 esclusi dai bottom-rank")

    st.markdown("---")
    st.markdown('<div class="section-title">🌳 Albero gerarchico (ordinato per Gross%)</div>', unsafe_allow_html=True)

    # Esporta gerarchia in CSV
    rows_export = []
    for am_name, am_data in tree.items():
        for dm_name, dm_data in am_data["dms"].items():
            for s in dm_data["stores"]:
                rows_export.append({
                    "Area Manager": am_name,
                    "District Manager": dm_name,
                    "Store": s["store"],
                    "Ragione Sociale": s["company"],
                    "Città": s["city"],
                    "Prov": s["province"],
                    "Tipo": s["type"],
                    "TAM": s["tam"],
                    "Gross": s["gross"],
                    "Gross %": safe_pct(s["gross"], s["tam"]),
                    "Net": s["net"],
                    "Net %": safe_pct(s["net"], s["tam"]),
                    "Forever": s["forever"],
                })
    if rows_export:
        export_df = pd.DataFrame(rows_export)
        csv_buf = export_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Esporta mappa zona (CSV)", data=csv_buf,
            file_name=f"mappa_zona_{zona_label}_{meta['month']:02d}_{meta['year']}.csv",
            mime="text/csv",
        )

    # Render albero
    for am_name, am_data in tree.items():
        st.markdown(f"""
        <div class="tree-am">
            <div class="tree-am-name">👔 {am_name}</div>
            <div style="font-size:0.78rem;color:#7B97BF;margin-top:0.2rem">
                {', '.join(am_data['regions']) or '—'} ·
                <b style="color:#E8F0FE">{am_data['n_dm']}</b> DM ·
                <b style="color:#E8F0FE">{am_data['n_stores']}</b> store ·
                Gross <b style="color:#00A8E0">{am_data['gross_pct']}%</b> ·
                Net <b style="color:#00A8E0">{am_data['net_pct']}%</b>
            </div>
        </div>
        """, unsafe_allow_html=True)

        for dm_name, dm_data in am_data["dms"].items():
            st.markdown(f"""
            <div class="tree-dm">
                <div class="tree-dm-name">🧑‍💼 {dm_name}</div>
                <div class="tree-dm-meta">
                    {', '.join(dm_data['regions']) or '—'} ·
                    <b style="color:#E8F0FE">{dm_data['n_stores']}</b> store ·
                    TAM {dm_data['tam']} ·
                    Gross <b style="color:#00A8E0">{dm_data['gross_pct']}%</b> ·
                    Net <b style="color:#00A8E0">{dm_data['net_pct']}%</b>
                </div>
            </div>
            """, unsafe_allow_html=True)
            if h_show_stores:
                for s in dm_data["stores"]:
                    g = s["gross_pct"]
                    n = s["net_pct"]
                    color_g = "#EF5350" if g < soglia_gross else ("#FFA726" if g < sa_gross else "#66BB6A")
                    color_n = "#EF5350" if n < soglia_net else ("#FFA726" if n < sa_net else "#66BB6A")
                    st.markdown(f"""
                    <div class="tree-store">
                        🏪 <b style="color:#E8F0FE">{s['store']}</b> · {s['city']} ({s['province']}) ·
                        <span style="color:#7B97BF">{s['type']}</span> ·
                        TAM {s['tam']} ·
                        Gross <span style="color:{color_g}"><b>{g}%</b></span> ·
                        Net <span style="color:{color_n}"><b>{n}%</b></span> ·
                        Forever {s['forever']}
                    </div>
                    """, unsafe_allow_html=True)


# ── TAB RAGIONI SOCIALI ─────────────────────────────────────────
with tab_rs:
    st.markdown('<div class="section-title">🏢 Ragioni Sociali della zona</div>', unsafe_allow_html=True)
    st.caption(
        f"Aggregazione di tutti i punti vendita per Ragione Sociale. "
        f"Totale: **{len(rs_df)} RS** sulla zona {zona_label}."
    )

    # Filtri
    col_rs1, col_rs2, col_rs3 = st.columns([2, 1, 1])
    with col_rs1:
        rs_search = st.text_input(
            "🔍 Cerca per nome", placeholder="Es: TELE RETAIL, BOOST AGENCY...",
            label_visibility="collapsed", key="rs_search"
        )
    with col_rs2:
        rs_tipo = st.selectbox(
            "Filtra tipo",
            ["Tutti"] + sorted(df["STORE_TYPE"].dropna().unique().tolist()),
            key="rs_tipo_filter",
        )
    with col_rs3:
        rs_sort = st.selectbox(
            "Ordina per",
            ["TOTALE Gross", "Gross %", "N PDV", "TAM", "Forever", "Nome"],
            key="rs_sort",
        )

    # Applica filtri
    rs_filtered = rs_df.copy()
    if rs_search.strip():
        q = rs_search.strip().upper()
        rs_filtered = rs_filtered[rs_filtered["COMPANY_NAME"].astype(str).str.upper().str.contains(q, na=False)]
    if rs_tipo != "Tutti":
        rs_filtered = rs_filtered[rs_filtered["Tipi_Store"].astype(str).str.contains(rs_tipo, na=False)]

    sort_map = {
        "TOTALE Gross": ("Gross", False),
        "Gross %": ("Gross_%", False),
        "N PDV": ("N_PDV", False),
        "TAM": ("TAM", False),
        "Forever": ("Forever_Active", False),
        "Nome": ("COMPANY_NAME", True),
    }
    sc, asc = sort_map[rs_sort]
    rs_filtered = rs_filtered.sort_values(sc, ascending=asc)

    st.markdown(f"**{len(rs_filtered)} ragioni sociali**")

    # Tabella
    rs_display_cols = {
        "COMPANY_NAME": "Ragione Sociale",
        "N_PDV": "N PDV",
        "Tipi_Store": "Tipo",
        "AM": "AM",
        "DM": "DM",
        "Regioni": "Regioni",
        "TAM": "TAM",
        "Gross": "Gross",
        "Gross_%": "Gross %",
        "Net": "Net",
        "Net_%": "Net %",
        "Forever_Active": "Forever",
    }
    rs_view = rs_filtered[list(rs_display_cols.keys())].rename(columns=rs_display_cols)

    def cg_rs(v):
        if v < soglia_gross: return "background-color:#FFCDD2;color:#B71C1C"
        if v < sa_gross:     return "background-color:#FFE0B2;color:#E65100"
        return "background-color:#C8E6C9;color:#1B5E20"
    def cn_rs(v):
        if v < soglia_net: return "background-color:#FFCDD2;color:#B71C1C"
        if v < sa_net:     return "background-color:#FFE0B2;color:#E65100"
        return "background-color:#C8E6C9;color:#1B5E20"

    st.dataframe(
        rs_view.style.map(cg_rs, subset=["Gross %"]).map(cn_rs, subset=["Net %"]),
        use_container_width=True, height=500,
    )

    # Esporta CSV
    rs_csv = rs_filtered.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Esporta ragioni sociali (CSV)", data=rs_csv,
        file_name=f"ragioni_sociali_{zona_label}_{meta['month']:02d}_{meta['year']}.csv",
        mime="text/csv",
    )

    # Drill-down singola RS
    st.markdown("---")
    st.markdown('<div class="section-title">🔍 Drill-down singola Ragione Sociale</div>', unsafe_allow_html=True)
    rs_list = rs_filtered["COMPANY_NAME"].tolist()
    if rs_list:
        rs_selected = st.selectbox(
            "Seleziona RS", rs_list, key="rs_drill",
            format_func=lambda x: f"{x[:60]}",
        )
        if rs_selected:
            stores_rs = df[df["COMPANY_NAME"] == rs_selected].copy()
            stores_rs["Gross_%"] = stores_rs.apply(
                lambda r: gross_pct_calc(r["SR_PLUS_GROSS_SALES"], r["TAM"], r["Activeforever"], calc_mode), axis=1)
            stores_rs["Net_%"] = stores_rs.apply(
                lambda r: net_pct_calc(r["SR_PLUS_NET_SALES"], r["TAM"], r["Activeforever"], calc_mode), axis=1)

            rs_row = rs_filtered[rs_filtered["COMPANY_NAME"] == rs_selected].iloc[0]
            col_d1, col_d2, col_d3, col_d4 = st.columns(4)
            col_d1.metric("PDV", int(rs_row["N_PDV"]))
            col_d2.metric("TAM totale", f"{int(rs_row['TAM']):,}")
            col_d3.metric("Gross / Net", f"{int(rs_row['Gross'])} / {int(rs_row['Net'])}")
            col_d4.metric("Forever attive", int(rs_row["Forever_Active"]))

            store_cols = ["STORE", "CITY", "PROVINCE_CODE", "STORE_TYPE",
                          "AREA_MANAGER", "DISTRICT_MANAGER",
                          "TAM", "SR_PLUS_GROSS_SALES", "Gross_%",
                          "SR_PLUS_NET_SALES", "Net_%", "Activeforever"]
            store_rename = {
                "STORE": "Store", "CITY": "Città", "PROVINCE_CODE": "Prov",
                "STORE_TYPE": "Tipo", "AREA_MANAGER": "AM", "DISTRICT_MANAGER": "DM",
                "TAM": "TAM", "SR_PLUS_GROSS_SALES": "Gross", "Gross_%": "Gross%",
                "SR_PLUS_NET_SALES": "Net", "Net_%": "Net%", "Activeforever": "Forever",
            }
            stores_view = stores_rs[store_cols].rename(columns=store_rename).sort_values("Gross%", ascending=False)
            st.dataframe(
                stores_view.style.map(cg_rs, subset=["Gross%"]).map(cn_rs, subset=["Net%"]),
                use_container_width=True,
            )


# ── TAB COMPENSI ────────────────────────────────────────────────
with tab_comp:
    st.markdown('<div class="section-title">💰 Compensi Reload stimati per Ragione Sociale</div>', unsafe_allow_html=True)
    st.caption(
        "Calcolo basato sui gettoni dello Smartphone Reload (PPTX Aprile 2026, slide 16/32). "
        "Include: Reload base per fascia di prezzo, Forever Basic/Premium, Reload Plus, Reload EU, Exchange. "
        "Il moltiplicatore X1/X2/X3/X4 è applicato in base al Net Attachment Rate per canale."
    )

    # Avviso importante
    st.markdown("""
    <div style="background:rgba(255,167,38,0.08);border-left:3px solid #FFA726;
        padding:0.7rem 1rem;border-radius:6px;font-size:0.82rem;margin:0.5rem 0;">
        ⚠️ <b>Stima orientativa.</b> I valori sono calcolati con i gettoni del PPTX
        Aprile 2026 e con le seguenti semplificazioni: Plus e EU usano un gettone medio
        (8.5€ e 9.5€) finché non avremo lo split per fascia. Il canale di ogni RS è
        determinato dal STORE_TYPE dominante dei suoi PDV.
    </div>
    """, unsafe_allow_html=True)

    # KPI di sintesi
    if not compensi_df.empty:
        tot_compenso = compensi_df["TOTALE stimato (€)"].sum()
        n_rs = len(compensi_df)
        media_rs = tot_compenso / n_rs if n_rs else 0
        max_rs = compensi_df["TOTALE stimato (€)"].max()
        n_x4 = (compensi_df["Moltipl."] == "x4").sum()
        n_x3 = (compensi_df["Moltipl."] == "x3").sum()
        n_x2 = (compensi_df["Moltipl."] == "x2").sum()
        n_x1 = (compensi_df["Moltipl."] == "x1").sum()

        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        with col_c1:
            st.markdown(f"""
            <div class="kpi-tile accent">
                <div class="kpi-label">💰 TOTALE zona</div>
                <div class="kpi-value accent">€ {tot_compenso:,.0f}</div>
                <div class="kpi-sub">{n_rs} ragioni sociali</div>
            </div>""", unsafe_allow_html=True)
        with col_c2:
            st.markdown(f"""
            <div class="kpi-tile green">
                <div class="kpi-label">📊 Media per RS</div>
                <div class="kpi-value green">€ {media_rs:,.0f}</div>
                <div class="kpi-sub">stima mensile</div>
            </div>""", unsafe_allow_html=True)
        with col_c3:
            st.markdown(f"""
            <div class="kpi-tile accent">
                <div class="kpi-label">🏆 Top RS</div>
                <div class="kpi-value accent">€ {max_rs:,.0f}</div>
                <div class="kpi-sub">migliore della zona</div>
            </div>""", unsafe_allow_html=True)
        with col_c4:
            st.markdown(f"""
            <div class="kpi-tile orange">
                <div class="kpi-label">🚀 Moltiplicatori</div>
                <div class="kpi-value" style="font-size:1.2rem;color:#E8F0FE">
                    x4:<b style="color:#66BB6A">{n_x4}</b> · x3:<b style="color:#66BB6A">{n_x3}</b><br>
                    x2:<b style="color:#FFA726">{n_x2}</b> · x1:<b style="color:#EF5350">{n_x1}</b>
                </div>
                <div class="kpi-sub">distribuzione AR</div>
            </div>""", unsafe_allow_html=True)

        st.markdown(" ")

        # Filtri
        col_cf1, col_cf2, col_cf3 = st.columns([2, 1, 1])
        with col_cf1:
            comp_search = st.text_input(
                "🔍 Cerca RS", placeholder="Nome ragione sociale...",
                key="comp_search", label_visibility="collapsed",
            )
        with col_cf2:
            comp_canale = st.selectbox(
                "Canale",
                ["Tutti"] + sorted(compensi_df["Canale"].unique().tolist()),
                key="comp_canale",
            )
        with col_cf3:
            comp_molt = st.selectbox(
                "Moltiplicatore",
                ["Tutti", "x4", "x3", "x2", "x1"],
                key="comp_molt",
            )

        comp_filtered = compensi_df.copy()
        if comp_search.strip():
            q = comp_search.strip().upper()
            comp_filtered = comp_filtered[
                comp_filtered["Ragione Sociale"].astype(str).str.upper().str.contains(q, na=False)
            ]
        if comp_canale != "Tutti":
            comp_filtered = comp_filtered[comp_filtered["Canale"] == comp_canale]
        if comp_molt != "Tutti":
            comp_filtered = comp_filtered[comp_filtered["Moltipl."] == comp_molt]

        st.markdown(f"**{len(comp_filtered)} RS** · Totale filtrato: € {comp_filtered['TOTALE stimato (€)'].sum():,.0f}")

        # Tabella compensi (escludo colonne tecniche dalla vista)
        display_compensi = comp_filtered.drop(
            columns=["_dettaglio_fasce", "_canale_int", "_net_ar_raw", "_subtotal_raw"],
            errors="ignore"
        )
        st.dataframe(display_compensi, use_container_width=True, height=500, hide_index=True)

        # Esporta CSV compensi
        comp_csv = display_compensi.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Esporta compensi (CSV)", data=comp_csv,
            file_name=f"compensi_reload_{zona_label}_{meta['month']:02d}_{meta['year']}.csv",
            mime="text/csv",
        )

        # Drill-down compenso singola RS
        st.markdown("---")
        st.markdown('<div class="section-title">🔬 Dettaglio compenso per RS</div>', unsafe_allow_html=True)
        if not comp_filtered.empty:
            rs_comp_sel = st.selectbox(
                "Seleziona RS", comp_filtered["Ragione Sociale"].tolist(), key="comp_drill",
            )
            if rs_comp_sel:
                row = comp_filtered[comp_filtered["Ragione Sociale"] == rs_comp_sel].iloc[0]
                col_dc1, col_dc2 = st.columns(2)
                with col_dc1:
                    st.markdown(f"**Canale:** {row['Canale']}")
                    st.markdown(f"**N PDV:** {int(row['N PDV'])}")
                    st.markdown(f"**TAM:** {int(row['TAM']):,}")
                    st.markdown(f"**Gross / Net:** {int(row['Gross'])} / {int(row['Net'])}")
                    st.markdown(f"**Forever attive:** {int(row['Forever'])}")
                    st.markdown(f"**Net AR:** {row['Net AR %']}%")
                    st.markdown(f"**Moltiplicatore:** **{row['Moltipl.']}**")
                with col_dc2:
                    st.markdown(f"**Compenso base reload:** € {row['Compenso base reload']:,.2f}")
                    st.markdown(f"**Forever Basic:** € {row['Compenso Forever Basic']:,.2f}")
                    st.markdown(f"**Forever Premium:** € {row['Compenso Forever Premium']:,.2f}")
                    st.markdown(f"**Reload Plus:** € {row['Compenso Plus']:,.2f}")
                    st.markdown(f"**Reload EU:** € {row['Compenso EU']:,.2f}")
                    st.markdown(f"**Exchange:** € {row['Compenso Exchange']:,.2f}")
                    st.markdown(f"**Subtotale:** € {row['Subtotale (€)']:,.2f}")
                    st.markdown(f"### 💰 TOTALE: € {row['TOTALE stimato (€)']:,.2f}")

                # Dettaglio per fascia
                if "_dettaglio_fasce" in row and isinstance(row["_dettaglio_fasce"], dict):
                    st.markdown("**Dettaglio per fascia di prezzo telefono:**")
                    fasce_rows = []
                    for fascia in FASCE_PPTX:
                        d = row["_dettaglio_fasce"].get(fascia["key"], {})
                        fasce_rows.append({
                            "Fascia": fascia["label"],
                            "TAM": d.get("tam", 0),
                            "Gross": d.get("gross", 0),
                            "Gettone €": d.get("gettone", 0),
                            "Compenso €": round(d.get("compenso", 0), 2),
                        })
                    st.dataframe(pd.DataFrame(fasce_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Nessun compenso calcolato.")

    # ─────────────────────────────────────────────────────────
    #  EXTRA GARA RELOAD FOREVER (Fase 3) — gara trimestrale
    # ─────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<div class="section-title">🚀 Extra Gara Reload Forever (trimestrale)</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Gara per Ragione Sociale (PPTX slide 17/33). Target 100/200/300 Forever per PDV → "
        "premi 500€ / 1.500€ / 3.000€ per PDV. I target sono moltiplicati per il numero di PDV "
        "della RS. **Periodo trimestrale** — il file mensile mostra solo il dato del mese corrente."
    )
    st.markdown("""
    <div style="background:rgba(0,168,224,0.06);border-left:3px solid #00A8E0;
        padding:0.7rem 1rem;border-radius:6px;font-size:0.82rem;margin:0.5rem 0;">
        ℹ️ Il calcolo usa le <b>Forever R4 Net</b> (Basic+Premium) del mese caricato. Per il
        dato trimestrale completo dovrai sommare 3 mesi (febbraio+marzo+aprile per l'edizione
        attuale). In una prossima iterazione aggiungerò il caricamento multi-mese.
    </div>
    """, unsafe_allow_html=True)

    if not extra_forever_df.empty:
        # KPI sintesi
        tot_premio_extra = extra_forever_df["Premio (€)"].sum()
        n_in_premio = (extra_forever_df["Premio (€)"] > 0).sum()
        n_scag1 = (extra_forever_df["Scaglione"] == 1).sum()
        n_scag2 = (extra_forever_df["Scaglione"] == 2).sum()
        n_scag3 = (extra_forever_df["Scaglione"] == 3).sum()

        col_ef1, col_ef2, col_ef3, col_ef4 = st.columns(4)
        with col_ef1:
            st.markdown(f"""
            <div class="kpi-tile accent">
                <div class="kpi-label">💰 Premio extra zona</div>
                <div class="kpi-value accent">€ {tot_premio_extra:,.0f}</div>
                <div class="kpi-sub">{n_in_premio} RS in premio</div>
            </div>""", unsafe_allow_html=True)
        with col_ef2:
            st.markdown(f"""
            <div class="kpi-tile green">
                <div class="kpi-label">🥉 1° scaglione</div>
                <div class="kpi-value green">{n_scag1}</div>
                <div class="kpi-sub">500€ × PDV</div>
            </div>""", unsafe_allow_html=True)
        with col_ef3:
            st.markdown(f"""
            <div class="kpi-tile green">
                <div class="kpi-label">🥈 2° scaglione</div>
                <div class="kpi-value green">{n_scag2}</div>
                <div class="kpi-sub">1.500€ × PDV</div>
            </div>""", unsafe_allow_html=True)
        with col_ef4:
            st.markdown(f"""
            <div class="kpi-tile green">
                <div class="kpi-label">🥇 3° scaglione</div>
                <div class="kpi-value green">{n_scag3}</div>
                <div class="kpi-sub">3.000€ × PDV</div>
            </div>""", unsafe_allow_html=True)

        st.markdown(" ")

        # Filtro: solo RS in premio o vicine
        col_efa, col_efb = st.columns([1, 1])
        with col_efa:
            ef_filter = st.selectbox(
                "Mostra",
                ["Tutte", "Solo in premio", "Solo a meno di 50 dal prossimo target", "Solo fuori premio"],
                key="ef_filter",
            )
        with col_efb:
            ef_search = st.text_input(
                "🔍 Cerca RS", key="ef_search", placeholder="Nome ragione sociale...",
                label_visibility="collapsed",
            )

        ef_view = extra_forever_df.copy()
        if ef_filter == "Solo in premio":
            ef_view = ef_view[ef_view["Premio (€)"] > 0]
        elif ef_filter == "Solo a meno di 50 dal prossimo target":
            ef_view = ef_view[(ef_view["Gap"] > 0) & (ef_view["Gap"] <= 50)]
        elif ef_filter == "Solo fuori premio":
            ef_view = ef_view[ef_view["Premio (€)"] == 0]
        if ef_search.strip():
            q = ef_search.strip().upper()
            ef_view = ef_view[ef_view["Ragione Sociale"].astype(str).str.upper().str.contains(q, na=False)]

        st.markdown(f"**{len(ef_view)} RS** · Totale premio: € {ef_view['Premio (€)'].sum():,.0f}")

        ef_display = ef_view.drop(columns=["_forever_periodo_raw"], errors="ignore")
        st.dataframe(ef_display, use_container_width=True, height=400, hide_index=True)

        ef_csv = ef_display.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Esporta extra Forever (CSV)", data=ef_csv,
            file_name=f"extra_forever_{zona_label}_{meta['month']:02d}_{meta['year']}.csv",
            mime="text/csv", key="dl_extra_forever",
        )

    # ─────────────────────────────────────────────────────────
    #  NEXT MILESTONE (Fase 5) — RS che con poco effort salgono
    # ─────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<div class="section-title">🎯 Opportunità — RS vicine al prossimo moltiplicatore</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "RS che con pochi reload aggiuntivi (entro +30% del Net attuale) potrebbero passare al "
        "moltiplicatore successivo (X1→X2, X2→X3, X3→X4). Ordinate per guadagno extra stimato."
    )

    if next_milestone_df is not None and not next_milestone_df.empty:
        # Filtro realistico/tutti
        col_nm1, col_nm2 = st.columns([1, 1])
        with col_nm1:
            nm_filter = st.selectbox(
                "Mostra",
                ["Solo opportunità realistiche (✅)", "Tutte"],
                key="nm_filter",
            )
        with col_nm2:
            nm_search = st.text_input(
                "🔍 Cerca RS", key="nm_search", placeholder="Nome ragione sociale...",
                label_visibility="collapsed",
            )

        nm_view = next_milestone_df.copy()
        if nm_filter.startswith("Solo opportunità"):
            nm_view = nm_view[nm_view["Realistico"] == "✅"]
        if nm_search.strip():
            q = nm_search.strip().upper()
            nm_view = nm_view[nm_view["Ragione Sociale"].astype(str).str.upper().str.contains(q, na=False)]

        if not nm_view.empty:
            tot_potenziale = nm_view["Guadagno extra (€)"].sum()
            st.markdown(
                f"**{len(nm_view)} RS** · Guadagno potenziale extra zona: "
                f"<b style='color:#66BB6A'>€ {tot_potenziale:,.0f}</b>",
                unsafe_allow_html=True,
            )

            nm_display = nm_view.drop(columns=["_ratio"], errors="ignore")
            st.dataframe(nm_display, use_container_width=True, height=400, hide_index=True)

            nm_csv = nm_display.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Esporta opportunità (CSV)", data=nm_csv,
                file_name=f"opportunita_{zona_label}_{meta['month']:02d}_{meta['year']}.csv",
                mime="text/csv", key="dl_next_milestone",
            )
        else:
            st.info("Nessuna RS in opportunità realistica con i filtri attuali.")
    else:
        st.info("Tutte le RS sono già al moltiplicatore massimo X4 — niente opportunità di upgrade.")


# ── TAB 4: DOWNLOAD ────────────────────────────────────────────
with tab4:
    st.markdown('<div class="section-title">⬇️ Download pacchetto completo</div>', unsafe_allow_html=True)
    col_d1, col_d2, col_d3 = st.columns(3)

    with col_d1:
        st.markdown("**📊 Riepilogo Excel**")
        st.caption("Tutti i DM con KPI, trend, alert — colorato per soglie")
        buf_rie = build_excel_riepilogo(kpis, meta, soglia_gross, soglia_net, sa_gross, sa_net, zona_label, regioni)
        st.download_button(
            "⬇️ riepilogo_generale.xlsx", data=buf_rie,
            file_name=f"riepilogo_{zona_label}_{meta['month']:02d}_{meta['year']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_d2:
        st.markdown("**📄 Riepilogo PDF**")
        st.caption("Versione stampabile — con trend mese precedente")
        pdf_bytes = build_pdf_riepilogo(kpis, meta, soglia_gross, soglia_net, zona_label, regioni)
        st.download_button(
            "⬇️ riepilogo_generale.pdf", data=pdf_bytes,
            file_name=f"riepilogo_{zona_label}_{meta['month']:02d}_{meta['year']}.pdf",
            mime="application/pdf", use_container_width=True,
        )

    with col_d3:
        st.markdown("**💬 Tutti i messaggi .txt**")
        st.caption("Un messaggio per ogni DM, pronti da copiare su WhatsApp")
        sep = "=" * 60 + "\n"
        all_msgs = sep.join([generate_message(row, meta) + "\n\n" for _, row in kpis.iterrows()])
        st.download_button(
            "⬇️ messaggi_whatsapp.txt", data=all_msgs.encode("utf-8"),
            file_name=f"messaggi_{zona_label}_{meta['month']:02d}_{meta['year']}.txt",
            mime="text/plain", use_container_width=True,
        )

    st.markdown("---")
    st.markdown("**📦 ZIP — Report chiusura mese per DM**")
    st.caption(
        "Un file Excel per ciascun District Manager con 3 fogli: "
        "Riepilogo KPI + alert + WhatsApp, Dettaglio Store completo (TAM/Gross/Net/Forever/Plus/Exchange), "
        "Compensi RS stimati."
    )
    if st.button("🗜️ Genera ZIP report DM", use_container_width=True, key="zip_dm"):
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for _idx, row in kpis.iterrows():
                dm_buf = build_excel_dm(
                    row, df, meta, soglia_gross, soglia_net, sa_gross, sa_net,
                    compensi_df=compensi_df, calc_mode=calc_mode,
                )
                safe = (row["DISTRICT_MANAGER"]
                        .replace("/", "_").replace("\\", "_").replace(" ", "_")[:40])
                reg = safe_str(row.get("REGION"), "").replace(" ", "")[:10]
                zf.writestr(f"{safe}_{reg}_{meta['month']:02d}_{meta['year']}.xlsx", dm_buf.read())
        zip_buf.seek(0)
        st.download_button(
            "⬇️ Scarica ZIP DM", data=zip_buf,
            file_name=f"wind3_reload_DM_{zona_label}_{meta['month']:02d}_{meta['year']}.zip",
            mime="application/zip", use_container_width=True, key="dl_zip_dm",
        )

    st.markdown("---")
    st.markdown("**📦 ZIP — Report chiusura mese per Area Manager**")
    st.caption(
        "Un file Excel per ciascun AM con 3 fogli: "
        "Riepilogo AM (KPI aggregati, critici/ok), "
        "Classifica DM (ordinata per Gross%, colorata), "
        "Tutti gli Store dell'AM."
    )
    if st.button("🗜️ Genera ZIP report AM", use_container_width=True, key="zip_am"):
        zip_buf_am = io.BytesIO()
        am_list = kpis["AREA_MANAGER"].dropna().unique().tolist()
        with zipfile.ZipFile(zip_buf_am, "w", zipfile.ZIP_DEFLATED) as zf:
            for am_n in am_list:
                am_buf = build_excel_am(
                    am_n, kpis, df, meta, soglia_gross, soglia_net, sa_gross, sa_net,
                    compensi_df=compensi_df, zona_label=zona_label,
                )
                safe = safe_str(am_n, "AM").replace("/", "_").replace("\\", "_").replace(" ", "_")[:40]
                zf.writestr(f"AM_{safe}_{meta['month']:02d}_{meta['year']}.xlsx", am_buf.read())
        zip_buf_am.seek(0)
        st.download_button(
            "⬇️ Scarica ZIP AM", data=zip_buf_am,
            file_name=f"wind3_reload_AM_{zona_label}_{meta['month']:02d}_{meta['year']}.zip",
            mime="application/zip", use_container_width=True, key="dl_zip_am",
        )

    st.markdown("---")
    st.markdown(f"""
    <div style="text-align:center;color:#7B97BF;font-size:0.75rem;padding:1rem;">
        Wind3 Reload Dashboard v4 · Ambassador {zona_label} · {mese_str}<br>
        Soglie — Gross critica: {soglia_gross}% / attenzione: {sa_gross}% ·
        Net critica: {soglia_net}% / attenzione: {sa_net}%
    </div>
    """, unsafe_allow_html=True)
